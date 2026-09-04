from __future__ import annotations

import hashlib
import json
import sqlite3
import zlib
from collections.abc import Callable
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from xml.etree import ElementTree
from zipfile import (
    ZIP_DEFLATED,
    ZIP_STORED,
    BadZipFile,
    LargeZipFile,
    ZipFile,
)

from fastapi import APIRouter, Depends, Header, Request, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from starlette.datastructures import UploadFile
from starlette.exceptions import HTTPException as StarletteHTTPException
from starlette.types import Message

from backend.app.core.config import Settings
from backend.app.core.database import transaction_immediate
from backend.app.features import business_attachments
from backend.app.features import procurement as base
from backend.app.features.api_common import (
    ApiError,
    ApiErrorRoute,
    idempotency_storage_key,
    restore_idempotent_response,
    save_idempotent_response,
)
from backend.app.features.auth import require_authenticated_session
from backend.app.features.inventory import format_quantity, line_value, parse_quantity

Clock = Callable[[], datetime]

_MAX_IMPORT_BYTES = 20 * 1024 * 1024
_MAX_MULTIPART_BYTES = _MAX_IMPORT_BYTES + 256 * 1024
_MAX_IMPORT_ROWS = 10_000
_MAX_XLSX_MEMBERS = 256
_MAX_XLSX_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
_MAX_XLSX_MEMBER_BYTES = 64 * 1024 * 1024
_MAX_XLSX_COMPRESSION_RATIO = 200
_IMPORT_CONFIRM_FIELDS = {"list_name", "expected_revision"}
_ORDER_UPDATE_FIELDS = {*base._ORDER_FIELDS, "expected_revision"}
_CANCEL_FIELDS = {"reason", "expected_revision"}
_ALLOCATION_FIELDS = {"purchase_order_line_id", "amount_cents"}
_PAYMENT_FIELDS = {
    "paid_on",
    "amount_cents",
    "payment_method",
    "reference_no",
    "allocations",
    "notes",
}
_INVOICE_FIELDS = {
    "invoice_no",
    "invoiced_on",
    "amount_cents",
    "allocations",
    "document_version_ids",
}
_QUOTE_FIELDS = {"title", "customer_company_id", "notes"}
_CUSTOMER_QUOTE_HEADERS = (
    "序号",
    "大类",
    "名称",
    "规格",
    "品牌",
    "型号",
    "数量",
    "单位",
    "报价单价（元）",
    "报价金额（元）",
)


def create_procurement_extensions_router(
    get_connection: Callable[..., sqlite3.Connection],
    get_settings: Callable[..., Settings],
    *,
    clock: Clock | None = None,
) -> APIRouter:
    router = APIRouter(route_class=ApiErrorRoute, tags=["procurement"])
    connection_dependency = Depends(get_connection)
    settings_dependency = Depends(get_settings)
    now = clock or _utc_now

    def require_session(
        request: Request,
        settings: Settings = settings_dependency,
    ) -> None:
        require_authenticated_session(request, settings.session_secret)

    authentication_dependency = Depends(require_session)

    @router.post(
        "/api/projects/{project_code}/procurement-imports/preview",
        status_code=status.HTTP_201_CREATED,
    )
    async def preview_procurement_import(
        project_code: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        key = base._validate_idempotency_key(idempotency_key)
        filename, content = await _read_xlsx_upload(request)
        digest = hashlib.sha256(content).hexdigest()
        request_hash = base._request_hash({"filename": filename, "sha256": digest})
        timestamp = base._timestamp(now)
        expires_at = (
            datetime.fromisoformat(timestamp) + timedelta(hours=24)
        ).isoformat()
        with transaction_immediate(connection):
            project = base._project(
                connection, base._normalize_project_path(project_code)
            )
            scope = _project_scope(
                "POST", str(project["project_code"]), "/procurement-imports/preview"
            )
            restored = restore_idempotent_response(
                connection, scope=scope, key=key, request_hash=request_hash
            )
            if restored is not None:
                return restored
            _require_active_project(project)
            _reject_confirmed_import(connection, int(project["id"]), digest)
        preview = _parse_import_workbook(content)
        with transaction_immediate(connection):
            project = base._project(
                connection, base._normalize_project_path(project_code)
            )
            scope = _project_scope(
                "POST", str(project["project_code"]), "/procurement-imports/preview"
            )
            restored = restore_idempotent_response(
                connection, scope=scope, key=key, request_hash=request_hash
            )
            if restored is not None:
                return restored
            _require_active_project(project)
            _reject_confirmed_import(connection, int(project["id"]), digest)
            existing = connection.execute(
                """
                SELECT * FROM procurement_imports
                WHERE project_id = ? AND sha256 = ? AND status = 'preview'
                """,
                (project["id"], digest),
            ).fetchone()
            if existing is not None:
                if str(existing["expires_at"]) <= timestamp:
                    connection.execute(
                        """
                        UPDATE procurement_imports
                        SET filename = ?, preview_json = ?, revision = revision + 1,
                            expires_at = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            filename,
                            json.dumps(preview, ensure_ascii=False, sort_keys=True),
                            expires_at,
                            timestamp,
                            existing["id"],
                        ),
                    )
                    existing = _import_row(
                        connection, int(existing["id"]), int(project["id"])
                    )
                    if existing is None:
                        raise sqlite3.DatabaseError(
                            "refreshed procurement import is missing"
                        )
                row = existing
            else:
                cursor = connection.execute(
                    """
                    INSERT INTO procurement_imports
                        (project_id, filename, sha256, preview_json, status,
                         revision, expires_at, confirmed_list_id, created_at, updated_at)
                    VALUES (?, ?, ?, ?, 'preview', 1, ?, NULL, ?, ?)
                    """,
                    (
                        project["id"],
                        filename,
                        digest,
                        json.dumps(preview, ensure_ascii=False, sort_keys=True),
                        expires_at,
                        timestamp,
                        timestamp,
                    ),
                )
                row = _import_row(
                    connection, base._last_insert_id(cursor), int(project["id"])
                )
                if row is None:
                    raise sqlite3.DatabaseError("created procurement import is missing")
            response = _import_response(row, str(project["project_code"]))
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_201_CREATED,
                resource_type="procurement_import",
                resource_id=int(row["id"]),
                created_at=timestamp,
            )
            return response

    @router.post(
        "/api/projects/{project_code}/procurement-imports/{import_id}/confirm",
        status_code=status.HTTP_201_CREATED,
    )
    async def confirm_procurement_import(
        project_code: str,
        import_id: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = base._parse_identifier(import_id)
        key = base._validate_idempotency_key(idempotency_key)
        payload = await base._read_json(
            request,
            _IMPORT_CONFIRM_FIELDS,
            "Invalid procurement import payload",
        )
        normalized = {
            "list_name": base._required_text(
                payload["list_name"], "Invalid procurement import payload"
            ),
            "expected_revision": base._positive_integer(
                payload["expected_revision"], "Invalid procurement import payload"
            ),
        }
        request_hash = base._request_hash(normalized)
        timestamp = base._timestamp(now)
        with transaction_immediate(connection):
            project = base._project(connection, base._normalize_project_path(project_code))
            scope = _project_scope(
                "POST",
                str(project["project_code"]),
                f"/procurement-imports/{identifier}/confirm",
            )
            restored = restore_idempotent_response(
                connection, scope=scope, key=key, request_hash=request_hash
            )
            if restored is not None:
                return restored
            _require_active_project(project)
            import_row = _import_row(connection, identifier, int(project["id"]))
            if import_row is None:
                raise base._not_found("Procurement import not found")
            if import_row["status"] != "preview":
                raise base._business_conflict(
                    "Procurement import cannot be confirmed",
                    "PROCUREMENT_IMPORT_NOT_CONFIRMABLE",
                )
            base._require_revision(import_row, int(normalized["expected_revision"]))
            if str(import_row["expires_at"]) <= timestamp:
                raise base._business_conflict(
                    "Procurement import expired", "PROCUREMENT_IMPORT_EXPIRED"
                )
            preview = _stored_preview(import_row)
            errors = preview["errors"]
            rows = preview["rows"]
            if errors:
                raise base._business_conflict(
                    "Procurement import has validation errors",
                    "PROCUREMENT_IMPORT_HAS_ERRORS",
                )
            if not rows:
                raise base._business_conflict(
                    "Procurement import is empty", "PROCUREMENT_IMPORT_EMPTY"
                )
            storage_key = idempotency_storage_key(scope, key)
            inventory_matches = _inventory_exact_matches(connection)
            list_cursor = connection.execute(
                """
                INSERT INTO procurement_lists
                    (project_id, name, notes, status, revision,
                     create_idempotency_key, create_request_hash,
                     created_at, updated_at)
                VALUES (?, ?, NULL, 'draft', 1, ?, ?, ?, ?)
                """,
                (
                    project["id"],
                    normalized["list_name"],
                    _derived_storage_key(storage_key, "list"),
                    request_hash,
                    timestamp,
                    timestamp,
                ),
            )
            list_id = base._last_insert_id(list_cursor)
            for index, line in enumerate(rows, start=1):
                normalized_line = base._normalize_line(line)
                inventory_item_id = inventory_matches.get(
                    _inventory_identity_from_line(normalized_line)
                )
                connection.execute(
                    """
                    INSERT INTO procurement_lines
                        (procurement_list_id, inventory_item_id, sequence_no,
                         category, name, specification, brand, model,
                         quantity_milli, unit, unit_cost_cents,
                         quoted_unit_price_cents, revision,
                         create_idempotency_key, create_request_hash,
                         created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?)
                    """,
                    (
                        list_id,
                        inventory_item_id,
                        normalized_line["sequence_no"],
                        normalized_line["category"],
                        normalized_line["name"],
                        normalized_line["specification"],
                        normalized_line["brand"],
                        normalized_line["model"],
                        normalized_line["quantity_milli"],
                        normalized_line["unit"],
                        normalized_line["unit_cost_cents"],
                        normalized_line["quoted_unit_price_cents"],
                        _derived_storage_key(storage_key, f"line:{index}"),
                        base._request_hash(line),
                        timestamp,
                        timestamp,
                    ),
                )
            connection.execute(
                """
                UPDATE procurement_imports
                SET status = 'confirmed', revision = revision + 1,
                    confirmed_list_id = ?, updated_at = ?
                WHERE id = ? AND revision = ?
                """,
                (list_id, timestamp, identifier, normalized["expected_revision"]),
            )
            updated_import = _import_row(connection, identifier, int(project["id"]))
            list_row = base._list_row(connection, list_id, int(project["id"]))
            if updated_import is None or list_row is None:
                raise sqlite3.DatabaseError("confirmed procurement import is missing")
            response = {
                "import": _import_response(updated_import, str(project["project_code"])),
                "procurement_list": base._list_detail(
                    connection, list_row, str(project["project_code"])
                ),
            }
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_201_CREATED,
                resource_type="procurement_import",
                resource_id=identifier,
                created_at=timestamp,
            )
            return response

    @router.put("/api/projects/{project_code}/purchase-orders/{order_id}")
    async def update_purchase_order(
        project_code: str,
        order_id: str,
        request: Request,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = base._parse_identifier(order_id)
        payload = await base._read_json(
            request, _ORDER_UPDATE_FIELDS, "Invalid purchase order payload"
        )
        expected_revision = base._positive_integer(
            payload["expected_revision"], "Invalid purchase order payload"
        )
        normalized = base._normalize_order(
            {field: payload[field] for field in base._ORDER_FIELDS}
        )
        timestamp = base._timestamp(now)
        try:
            with transaction_immediate(connection):
                project = base._active_project(
                    connection, base._normalize_project_path(project_code)
                )
                order = base._order_row(connection, identifier, int(project["id"]))
                if order is None:
                    raise base._not_found("Purchase order not found")
                if order["status"] != "draft":
                    raise base._business_conflict(
                        "Purchase order is not editable", "PURCHASE_ORDER_NOT_EDITABLE"
                    )
                base._require_revision(order, expected_revision)
                _validate_order_references(connection, project, normalized)
                connection.execute(
                    """
                    UPDATE purchase_orders
                    SET order_no = ?, supplier_company_id = ?, ordered_on = ?,
                        expected_delivery_on = ?, notes = ?, revision = revision + 1,
                        updated_at = ?
                    WHERE id = ? AND revision = ?
                    """,
                    (
                        normalized["order_no"],
                        normalized["supplier_company_id"],
                        normalized["ordered_on"],
                        normalized["expected_delivery_on"],
                        normalized["notes"],
                        timestamp,
                        identifier,
                        expected_revision,
                    ),
                )
                connection.execute(
                    "DELETE FROM purchase_order_documents WHERE purchase_order_id = ?",
                    (identifier,),
                )
                connection.execute(
                    "DELETE FROM purchase_order_lines WHERE purchase_order_id = ?",
                    (identifier,),
                )
                _insert_order_lines(connection, identifier, normalized["lines"], timestamp)
                for document_id in normalized["document_version_ids"]:
                    connection.execute(
                        """
                        INSERT INTO purchase_order_documents
                            (purchase_order_id, document_version_id)
                        VALUES (?, ?)
                        """,
                        (identifier, document_id),
                    )
                updated = base._order_row(connection, identifier, int(project["id"]))
                if updated is None:
                    raise sqlite3.DatabaseError("updated purchase order is missing")
                return base._order_response(
                    connection, updated, str(project["project_code"])
                )
        except sqlite3.IntegrityError as exc:
            if base._is_unique_constraint(exc):
                raise base._business_conflict(
                    "Purchase order already exists", "PURCHASE_ORDER_EXISTS"
                ) from None
            raise

    @router.post("/api/projects/{project_code}/purchase-orders/{order_id}/cancel")
    async def cancel_purchase_order(
        project_code: str,
        order_id: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = base._parse_identifier(order_id)
        key = base._validate_idempotency_key(idempotency_key)
        payload = await base._read_json(
            request, _CANCEL_FIELDS, "Invalid purchase order cancellation payload"
        )
        normalized = {
            "reason": base._required_text(
                payload["reason"], "Invalid purchase order cancellation payload"
            ),
            "expected_revision": base._positive_integer(
                payload["expected_revision"],
                "Invalid purchase order cancellation payload",
            ),
        }
        request_hash = base._request_hash(normalized)
        timestamp = base._timestamp(now)
        with transaction_immediate(connection):
            project = base._project(connection, base._normalize_project_path(project_code))
            scope = _project_scope(
                "POST",
                str(project["project_code"]),
                f"/purchase-orders/{identifier}/cancel",
            )
            restored = restore_idempotent_response(
                connection, scope=scope, key=key, request_hash=request_hash
            )
            if restored is not None:
                return restored
            _require_active_project(project)
            order = base._order_row(connection, identifier, int(project["id"]))
            if order is None:
                raise base._not_found("Purchase order not found")
            if order["status"] not in {"draft", "confirmed"}:
                raise base._business_conflict(
                    "Purchase order cannot be cancelled",
                    "PURCHASE_ORDER_NOT_CANCELLABLE",
                )
            base._require_revision(order, int(normalized["expected_revision"]))
            if _order_has_active_facts(connection, identifier):
                raise base._business_conflict(
                    "Purchase order has active business records",
                    "PURCHASE_ORDER_HAS_ACTIVE_FACTS",
                )
            connection.execute(
                """
                UPDATE purchase_orders
                SET status = 'cancelled', cancelled_at = ?, cancel_reason = ?,
                    revision = revision + 1, updated_at = ?
                WHERE id = ? AND revision = ?
                """,
                (
                    timestamp,
                    normalized["reason"],
                    timestamp,
                    identifier,
                    normalized["expected_revision"],
                ),
            )
            cancelled = base._order_row(connection, identifier, int(project["id"]))
            if cancelled is None:
                raise sqlite3.DatabaseError("cancelled purchase order is missing")
            response = base._order_response(
                connection, cancelled, str(project["project_code"])
            )
            response["cancelled_at"] = cancelled["cancelled_at"]
            response["cancel_reason"] = cancelled["cancel_reason"]
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_200_OK,
                resource_type="purchase_order",
                resource_id=identifier,
                created_at=timestamp,
            )
            return response

    @router.post(
        "/api/projects/{project_code}/purchase-orders/{order_id}/supplier-payments",
        status_code=status.HTTP_201_CREATED,
    )
    async def create_supplier_payment(
        project_code: str,
        order_id: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = base._parse_identifier(order_id)
        key = base._validate_idempotency_key(idempotency_key)
        payload = await base._read_json(
            request, _PAYMENT_FIELDS, "Invalid supplier payment payload"
        )
        normalized = _normalize_payment(payload)
        request_hash = base._request_hash(normalized)
        timestamp = base._timestamp(now)
        with transaction_immediate(connection):
            project = base._project(connection, base._normalize_project_path(project_code))
            scope = _project_scope(
                "POST",
                str(project["project_code"]),
                f"/purchase-orders/{identifier}/supplier-payments",
            )
            restored = restore_idempotent_response(
                connection, scope=scope, key=key, request_hash=request_hash
            )
            if restored is not None:
                return restored
            _require_active_project(project)
            order = _payable_order(connection, identifier, int(project["id"]))
            _validate_allocation_capacity(
                connection,
                identifier,
                normalized["allocations"],
                fact="payment",
            )
            storage_key = idempotency_storage_key(scope, key)
            cursor = connection.execute(
                """
                INSERT INTO supplier_payments
                    (purchase_order_id, paid_on, amount_cents, payment_method,
                     reference_no, notes, status, revision, idempotency_key,
                     request_hash, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, 'active', 1, ?, ?, ?, ?)
                """,
                (
                    order["id"],
                    normalized["paid_on"],
                    normalized["amount_cents"],
                    normalized["payment_method"],
                    normalized["reference_no"],
                    normalized["notes"],
                    storage_key,
                    request_hash,
                    timestamp,
                    timestamp,
                ),
            )
            payment_id = base._last_insert_id(cursor)
            for allocation in normalized["allocations"]:
                connection.execute(
                    """
                    INSERT INTO supplier_payment_allocations
                        (supplier_payment_id, purchase_order_line_id, amount_cents)
                    VALUES (?, ?, ?)
                    """,
                    (
                        payment_id,
                        allocation["purchase_order_line_id"],
                        allocation["amount_cents"],
                    ),
                )
            payment = _payment_row(connection, payment_id, int(project["id"]))
            if payment is None:
                raise sqlite3.DatabaseError("created supplier payment is missing")
            response = base._payment_response(connection, payment)
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_201_CREATED,
                resource_type="supplier_payment",
                resource_id=payment_id,
                created_at=timestamp,
            )
            return response

    @router.post(
        "/api/projects/{project_code}/supplier-payments/{payment_id}/reverse"
    )
    async def reverse_supplier_payment(
        project_code: str,
        payment_id: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = base._parse_identifier(payment_id)
        key, normalized = await _reversal_request(request, idempotency_key)
        request_hash = base._request_hash(normalized)
        timestamp = base._timestamp(now)
        with transaction_immediate(connection):
            project = base._project(connection, base._normalize_project_path(project_code))
            scope = _project_scope(
                "POST",
                str(project["project_code"]),
                f"/supplier-payments/{identifier}/reverse",
            )
            restored = restore_idempotent_response(
                connection, scope=scope, key=key, request_hash=request_hash
            )
            if restored is not None:
                return restored
            _require_active_project(project)
            payment = _payment_row(connection, identifier, int(project["id"]))
            if payment is None:
                raise base._not_found("Supplier payment not found")
            if payment["status"] != "active":
                raise base._business_conflict(
                    "Supplier payment is already reversed",
                    "SUPPLIER_PAYMENT_ALREADY_REVERSED",
                )
            base._require_revision(payment, int(normalized["expected_revision"]))
            connection.execute(
                """
                UPDATE supplier_payments
                SET status = 'reversed', reversal_reason = ?, reversed_at = ?,
                    revision = revision + 1, updated_at = ?
                WHERE id = ? AND revision = ?
                """,
                (
                    normalized["reason"],
                    timestamp,
                    timestamp,
                    identifier,
                    normalized["expected_revision"],
                ),
            )
            reversed_row = _payment_row(connection, identifier, int(project["id"]))
            if reversed_row is None:
                raise sqlite3.DatabaseError("reversed supplier payment is missing")
            response = base._payment_response(connection, reversed_row)
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_200_OK,
                resource_type="supplier_payment",
                resource_id=identifier,
                created_at=timestamp,
            )
            return response

    @router.post(
        "/api/projects/{project_code}/purchase-orders/{order_id}/supplier-invoices",
        status_code=status.HTTP_201_CREATED,
    )
    async def create_supplier_invoice(
        project_code: str,
        order_id: str,
        request: Request,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
        settings: Settings = settings_dependency,
    ) -> dict[str, object]:
        identifier = base._parse_identifier(order_id)
        key = _read_idempotency_key(request)
        multipart = business_attachments.is_multipart_request(request)
        if multipart:
            payload, attachment_batch = (
                await business_attachments.read_multipart_batch(
                    request,
                    data_dir=settings.data_dir,
                    max_file_size_bytes=(
                        settings.max_document_upload_mb * 1024 * 1024
                    ),
                    invalid=_invoice_attachment_error,
                    too_large=_attachment_too_large,
                    batch_too_large=_attachment_batch_too_large,
                )
            )
        else:
            payload = await base._read_json(
                request, _INVOICE_FIELDS, "Invalid supplier invoice payload"
            )
            attachment_batch = business_attachments.AttachmentBatch(
                [], settings.data_dir
            )
        try:
            with attachment_batch:
                if set(payload) != _INVOICE_FIELDS:
                    raise base._invalid_payload("Invalid supplier invoice payload")
                normalized = _normalize_invoice(payload)
                request_hash = base._request_hash(
                    attachment_batch.hash_payload(normalized)
                )
                timestamp = base._timestamp(now)
                with transaction_immediate(connection):
                    project = base._project(
                        connection, base._normalize_project_path(project_code)
                    )
                    scope = _project_scope(
                        "POST",
                        str(project["project_code"]),
                        f"/purchase-orders/{identifier}/supplier-invoices",
                    )
                    restored = restore_idempotent_response(
                        connection, scope=scope, key=key, request_hash=request_hash
                    )
                    if restored is not None:
                        return restored
                    _require_active_project(project)
                    order = _payable_order(
                        connection, identifier, int(project["id"])
                    )
                    base._validate_document_versions(
                        connection,
                        str(project["project_code"]),
                        normalized["document_version_ids"],
                    )
                    _validate_allocation_capacity(
                        connection,
                        identifier,
                        normalized["allocations"],
                        fact="invoice",
                    )
                    storage_key = idempotency_storage_key(scope, key)
                    cursor = connection.execute(
                        """
                        INSERT INTO supplier_invoices
                            (purchase_order_id, invoice_no, invoiced_on, amount_cents,
                             status, revision, idempotency_key, request_hash,
                             created_at, updated_at)
                        VALUES (?, ?, ?, ?, 'active', 1, ?, ?, ?, ?)
                        """,
                        (
                            order["id"],
                            normalized["invoice_no"],
                            normalized["invoiced_on"],
                            normalized["amount_cents"],
                            storage_key,
                            request_hash,
                            timestamp,
                            timestamp,
                        ),
                    )
                    invoice_id = base._last_insert_id(cursor)
                    for allocation in normalized["allocations"]:
                        connection.execute(
                            """
                            INSERT INTO supplier_invoice_allocations
                                (supplier_invoice_id, purchase_order_line_id, amount_cents)
                            VALUES (?, ?, ?)
                            """,
                            (
                                invoice_id,
                                allocation["purchase_order_line_id"],
                                allocation["amount_cents"],
                            ),
                        )
                    uploaded_version_ids = attachment_batch.publish_documents(
                        connection,
                        project_code=str(project["project_code"]),
                        category="invoice",
                        documents=[
                            business_attachments.ManagedDocument(
                                title=(
                                    f"进项发票 {normalized['invoice_no']}"
                                    f" 附件 {index:02d}（发票记录 {invoice_id}）"
                                ),
                                managed_filename=(
                                    business_attachments.managed_filename(
                                        project["project_code"],
                                        "进项发票",
                                        normalized["invoice_no"],
                                        f"{index:02d}",
                                        original_filename=(
                                            attachment.original_filename
                                        ),
                                    )
                                ),
                            )
                            for index, attachment in enumerate(
                                attachment_batch.attachments, start=1
                            )
                        ],
                        notes=None,
                        timestamp=timestamp,
                    )
                    document_version_ids = [
                        *normalized["document_version_ids"],
                        *uploaded_version_ids,
                    ]
                    for document_id in document_version_ids:
                        connection.execute(
                            """
                            INSERT INTO supplier_invoice_documents
                                (supplier_invoice_id, document_version_id)
                            VALUES (?, ?)
                            """,
                            (invoice_id, document_id),
                        )
                    invoice = _invoice_row(
                        connection, invoice_id, int(project["id"])
                    )
                    if invoice is None:
                        raise sqlite3.DatabaseError(
                            "created supplier invoice is missing"
                        )
                    response = base._invoice_response(connection, invoice)
                    save_idempotent_response(
                        connection,
                        scope=scope,
                        key=key,
                        request_hash=request_hash,
                        response=response,
                        response_status=status.HTTP_201_CREATED,
                        resource_type="supplier_invoice",
                        resource_id=invoice_id,
                        created_at=timestamp,
                    )
                    return response
        except sqlite3.IntegrityError as exc:
            if base._is_unique_constraint(exc):
                raise base._business_conflict(
                    "Supplier invoice already exists", "SUPPLIER_INVOICE_EXISTS"
                ) from None
            raise

    @router.post(
        "/api/projects/{project_code}/supplier-invoices/{invoice_id}/reverse"
    )
    async def reverse_supplier_invoice(
        project_code: str,
        invoice_id: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = base._parse_identifier(invoice_id)
        key, normalized = await _reversal_request(request, idempotency_key)
        request_hash = base._request_hash(normalized)
        timestamp = base._timestamp(now)
        with transaction_immediate(connection):
            project = base._project(connection, base._normalize_project_path(project_code))
            scope = _project_scope(
                "POST",
                str(project["project_code"]),
                f"/supplier-invoices/{identifier}/reverse",
            )
            restored = restore_idempotent_response(
                connection, scope=scope, key=key, request_hash=request_hash
            )
            if restored is not None:
                return restored
            _require_active_project(project)
            invoice = _invoice_row(connection, identifier, int(project["id"]))
            if invoice is None:
                raise base._not_found("Supplier invoice not found")
            if invoice["status"] != "active":
                raise base._business_conflict(
                    "Supplier invoice is already reversed",
                    "SUPPLIER_INVOICE_ALREADY_REVERSED",
                )
            base._require_revision(invoice, int(normalized["expected_revision"]))
            connection.execute(
                """
                UPDATE supplier_invoices
                SET status = 'reversed', reversal_reason = ?, reversed_at = ?,
                    revision = revision + 1, updated_at = ?
                WHERE id = ? AND revision = ?
                """,
                (
                    normalized["reason"],
                    timestamp,
                    timestamp,
                    identifier,
                    normalized["expected_revision"],
                ),
            )
            reversed_row = _invoice_row(connection, identifier, int(project["id"]))
            if reversed_row is None:
                raise sqlite3.DatabaseError("reversed supplier invoice is missing")
            response = base._invoice_response(connection, reversed_row)
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_200_OK,
                resource_type="supplier_invoice",
                resource_id=identifier,
                created_at=timestamp,
            )
            return response

    @router.post(
        "/api/projects/{project_code}/goods-receipts/{receipt_id}/reverse"
    )
    async def reverse_goods_receipt(
        project_code: str,
        receipt_id: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = base._parse_identifier(receipt_id)
        key, normalized = await _reversal_request(request, idempotency_key)
        request_hash = base._request_hash(normalized)
        timestamp = base._timestamp(now)
        with transaction_immediate(connection):
            project = base._project(connection, base._normalize_project_path(project_code))
            scope = _project_scope(
                "POST",
                str(project["project_code"]),
                f"/goods-receipts/{identifier}/reverse",
            )
            restored = restore_idempotent_response(
                connection, scope=scope, key=key, request_hash=request_hash
            )
            if restored is not None:
                return restored
            _require_active_project(project)
            receipt = _receipt_row(connection, identifier, int(project["id"]))
            if receipt is None:
                raise base._not_found("Goods receipt not found")
            if receipt["status"] != "active":
                raise base._business_conflict(
                    "Goods receipt is already reversed",
                    "GOODS_RECEIPT_ALREADY_REVERSED",
                )
            base._require_revision(receipt, int(normalized["expected_revision"]))
            _reverse_receipt_inventory(
                connection,
                receipt,
                int(project["id"]),
                str(normalized["reason"]),
                timestamp,
            )
            connection.execute(
                """
                UPDATE goods_receipts
                SET status = 'reversed', reversal_reason = ?, reversed_at = ?,
                    revision = revision + 1, updated_at = ?
                WHERE id = ? AND revision = ?
                """,
                (
                    normalized["reason"],
                    timestamp,
                    timestamp,
                    identifier,
                    normalized["expected_revision"],
                ),
            )
            _refresh_order_status_after_reversal(
                connection, int(receipt["purchase_order_id"]), timestamp
            )
            reversed_row = _receipt_row(connection, identifier, int(project["id"]))
            if reversed_row is None:
                raise sqlite3.DatabaseError("reversed goods receipt is missing")
            response = base._receipt_response(connection, reversed_row)
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_200_OK,
                resource_type="goods_receipt",
                resource_id=identifier,
                created_at=timestamp,
            )
            return response

    @router.post(
        "/api/projects/{project_code}/procurement-lists/{list_id}/quote-exports",
        status_code=status.HTTP_201_CREATED,
    )
    async def create_quote_export(
        project_code: str,
        list_id: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = base._parse_identifier(list_id)
        key = base._validate_idempotency_key(idempotency_key)
        payload = await base._read_json(
            request, _QUOTE_FIELDS, "Invalid quote export payload"
        )
        normalized = {
            "title": base._required_text(payload["title"], "Invalid quote export payload"),
            "customer_company_id": base._positive_integer(
                payload["customer_company_id"], "Invalid quote export payload"
            ),
            "notes": base._optional_text(payload["notes"], "Invalid quote export payload"),
        }
        request_hash = base._request_hash(normalized)
        timestamp = base._timestamp(now)
        with transaction_immediate(connection):
            project = base._project(connection, base._normalize_project_path(project_code))
            scope = _project_scope(
                "POST",
                str(project["project_code"]),
                f"/procurement-lists/{identifier}/quote-exports",
            )
            restored = restore_idempotent_response(
                connection, scope=scope, key=key, request_hash=request_hash
            )
            if restored is not None:
                return restored
            _require_active_project(project)
            if int(project["company_id"]) != normalized["customer_company_id"]:
                raise ApiError(
                    status.HTTP_409_CONFLICT,
                    "Quote customer must match project customer",
                    "PROJECT_CUSTOMER_MISMATCH",
                    field_errors={
                        "customer_company_id": ["必须使用项目绑定的客户公司"]
                    },
                    headers={"X-Error-Code": "PROJECT_CUSTOMER_MISMATCH"},
                )
            procurement_list = base._list_row(connection, identifier, int(project["id"]))
            if procurement_list is None:
                raise base._not_found("Procurement list not found")
            if procurement_list["status"] != "confirmed":
                raise base._business_conflict(
                    "Only a confirmed procurement list can be quoted",
                    "PROCUREMENT_LIST_NOT_QUOTABLE",
                )
            if connection.execute(
                "SELECT 1 FROM companies WHERE id = ?",
                (normalized["customer_company_id"],),
            ).fetchone() is None:
                raise base._not_found("Company not found")
            storage_key = idempotency_storage_key(scope, key)
            cursor = connection.execute(
                """
                INSERT INTO quote_exports
                    (project_id, procurement_list_id, title, customer_company_id,
                     notes, document_version_id, idempotency_key, request_hash,
                     created_at)
                VALUES (?, ?, ?, ?, ?, NULL, ?, ?, ?)
                """,
                (
                    project["id"],
                    identifier,
                    normalized["title"],
                    normalized["customer_company_id"],
                    normalized["notes"],
                    storage_key,
                    request_hash,
                    timestamp,
                ),
            )
            export_id = base._last_insert_id(cursor)
            export = _quote_row(connection, export_id, int(project["id"]))
            if export is None:
                raise sqlite3.DatabaseError("created quote export is missing")
            response = _quote_response(export, str(project["project_code"]))
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_201_CREATED,
                resource_type="quote_export",
                resource_id=export_id,
                created_at=timestamp,
            )
            return response

    @router.get("/api/projects/{project_code}/quote-exports")
    def list_quote_exports(
        project_code: str,
        request: Request,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        project = base._project(connection, base._normalize_project_path(project_code))
        page, page_size = base._read_pagination(request)
        total = int(
            connection.execute(
                "SELECT COUNT(*) FROM quote_exports WHERE project_id = ?",
                (project["id"],),
            ).fetchone()[0]
        )
        rows = connection.execute(
            """
            SELECT exports.*, companies.name AS customer_company_name
            FROM quote_exports AS exports
            JOIN companies ON companies.id = exports.customer_company_id
            WHERE exports.project_id = ?
            ORDER BY exports.created_at DESC, exports.id DESC
            LIMIT ? OFFSET ?
            """,
            (project["id"], page_size, (page - 1) * page_size),
        ).fetchall()
        normalized_code = str(project["project_code"])
        return {
            "items": [_quote_response(row, normalized_code) for row in rows],
            "page": page,
            "page_size": page_size,
            "total": total,
        }

    @router.get("/api/projects/{project_code}/quote-exports/{export_id}/download")
    def download_quote_export(
        project_code: str,
        export_id: str,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> StreamingResponse:
        project = base._project(connection, base._normalize_project_path(project_code))
        export = _quote_row(
            connection, base._parse_identifier(export_id), int(project["id"])
        )
        if export is None:
            raise base._not_found("Quote export not found")
        content = _quote_workbook(connection, export)
        return StreamingResponse(
            BytesIO(content),
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": (
                    f'attachment; filename="quote-export-{export["id"]}.xlsx"'
                )
            },
        )

    return router


class _RequestBodyTooLarge(Exception):
    pass


async def _read_xlsx_upload(request: Request) -> tuple[str, bytes]:
    received_bytes = 0

    async def limited_receive() -> Message:
        nonlocal received_bytes
        message = await request.receive()
        if message["type"] == "http.request":
            received_bytes += len(message.get("body", b""))
            if received_bytes > _MAX_MULTIPART_BYTES:
                raise _RequestBodyTooLarge
        return message

    limited_request = Request(request.scope, receive=limited_receive)
    try:
        async with limited_request.form(
            max_files=4,
            max_fields=4,
            max_part_size=_MAX_IMPORT_BYTES,
        ) as form:
            items = list(form.multi_items())
            if (
                len(items) != 1
                or items[0][0] != "file"
                or not isinstance(items[0][1], UploadFile)
            ):
                raise base._invalid_payload("Invalid procurement import file")
            upload = items[0][1]
            filename = upload.filename
            if not isinstance(filename, str) or not filename.casefold().endswith(
                ".xlsx"
            ):
                raise base._invalid_payload("Invalid procurement import file")
            content = await upload.read(_MAX_IMPORT_BYTES + 1)
    except ApiError:
        raise
    except (
        _RequestBodyTooLarge,
        RuntimeError,
        StarletteHTTPException,
        TypeError,
        ValueError,
    ):
        raise base._invalid_payload("Invalid procurement import file") from None
    if not content or len(content) > _MAX_IMPORT_BYTES:
        raise base._invalid_payload("Invalid procurement import file")
    return filename, content


def _parse_import_workbook(content: bytes) -> dict[str, list[dict[str, object]]]:
    _validate_xlsx_archive(content)
    workbook: Any | None = None
    try:
        workbook = load_workbook(
            BytesIO(content), read_only=True, data_only=False, keep_links=False
        )
        if not workbook.worksheets:
            raise base._invalid_payload("Invalid procurement import workbook")
        worksheet = workbook.worksheets[0]
        headers = tuple(cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1)))
        if headers != base._TEMPLATE_HEADERS:
            raise base._invalid_payload("Invalid procurement import headers")
        if worksheet.max_row > _MAX_IMPORT_ROWS + 1:
            raise base._invalid_payload("Procurement import exceeds 10000 rows")
        rows: list[dict[str, object]] = []
        errors: list[dict[str, object]] = []
        seen_sequences: set[int] = set()
        seen_business_rows: set[tuple[object, ...]] = set()
        data_row_count = 0
        for row_number, worksheet_row in enumerate(
            worksheet.iter_rows(min_row=2, max_col=len(base._TEMPLATE_HEADERS)),
            start=2,
        ):
            if all(cell.value is None for cell in worksheet_row):
                continue
            data_row_count += 1
            if data_row_count > _MAX_IMPORT_ROWS:
                raise base._invalid_payload("Procurement import exceeds 10000 rows")
            normalized, row_errors = _normalize_import_row(worksheet_row, row_number)
            sequence = normalized.get("sequence_no")
            if isinstance(sequence, int):
                if sequence in seen_sequences:
                    row_errors.append(
                        _cell_error(row_number, 1, "sequence_no", "序号重复")
                    )
                else:
                    seen_sequences.add(sequence)
            if not row_errors:
                fingerprint = tuple(
                    value
                    for field, value in normalized.items()
                    if field != "sequence_no"
                )
                if fingerprint in seen_business_rows:
                    row_errors.append(
                        _cell_error(row_number, 0, "row", "业务字段重复")
                    )
                else:
                    seen_business_rows.add(fingerprint)
            if row_errors:
                errors.extend(row_errors)
            else:
                rows.append(normalized)
        return {"rows": rows, "errors": errors}
    except ApiError:
        raise
    except (
        AttributeError,
        BadZipFile,
        ElementTree.ParseError,
        EOFError,
        InvalidFileException,
        IndexError,
        KeyError,
        LargeZipFile,
        OSError,
        OverflowError,
        StopIteration,
        TypeError,
        UnicodeError,
        ValueError,
        zlib.error,
    ):
        raise base._invalid_payload("Invalid procurement import workbook") from None
    finally:
        if workbook is not None:
            workbook.close()


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if not members or len(members) > _MAX_XLSX_MEMBERS:
                raise base._invalid_payload("Invalid procurement import workbook")
            names: set[str] = set()
            total_uncompressed = 0
            for member in members:
                normalized_parts = member.filename.replace("\\", "/").split("/")
                if (
                    member.filename in names
                    or member.filename.startswith(("/", "\\"))
                    or ".." in normalized_parts
                    or member.flag_bits & 0x1
                    or member.compress_type not in {ZIP_STORED, ZIP_DEFLATED}
                    or member.file_size > _MAX_XLSX_MEMBER_BYTES
                    or (
                        member.file_size > 0
                        and (
                            member.compress_size == 0
                            or member.file_size
                            > member.compress_size * _MAX_XLSX_COMPRESSION_RATIO
                        )
                    )
                ):
                    raise base._invalid_payload("Invalid procurement import workbook")
                names.add(member.filename)
                total_uncompressed += member.file_size
                if total_uncompressed > _MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise base._invalid_payload("Invalid procurement import workbook")
            if archive.testzip() is not None:
                raise base._invalid_payload("Invalid procurement import workbook")
    except ApiError:
        raise
    except (
        BadZipFile,
        EOFError,
        LargeZipFile,
        OSError,
        OverflowError,
        ValueError,
        zlib.error,
    ):
        raise base._invalid_payload("Invalid procurement import workbook") from None


def _reject_confirmed_import(
    connection: sqlite3.Connection,
    project_id: int,
    digest: str,
) -> None:
    confirmed = connection.execute(
        """
        SELECT 1 FROM procurement_imports
        WHERE project_id = ? AND sha256 = ? AND status = 'confirmed'
        """,
        (project_id, digest),
    ).fetchone()
    if confirmed is not None:
        raise base._business_conflict(
            "Procurement import was already confirmed",
            "PROCUREMENT_IMPORT_ALREADY_CONFIRMED",
        )


def _normalize_import_row(
    cells: tuple[Any, ...],
    row_number: int,
) -> tuple[dict[str, object], list[dict[str, object]]]:
    fields = (
        "sequence_no",
        "category",
        "name",
        "specification",
        "brand",
        "model",
        "quantity",
        "unit",
        "unit_cost_cents",
        "quoted_unit_price_cents",
    )
    normalized: dict[str, object] = {}
    errors: list[dict[str, object]] = []
    for column, (field, cell) in enumerate(zip(fields, cells, strict=True), start=1):
        try:
            if cell.data_type == "f":
                raise ValueError
            if field == "sequence_no":
                normalized[field] = _spreadsheet_positive_integer(cell.value)
            elif field in {"category", "name", "unit"}:
                normalized[field] = base._required_text(
                    cell.value, "Invalid procurement import cell"
                )
            elif field in {"specification", "brand", "model"}:
                normalized[field] = base._optional_text(
                    cell.value, "Invalid procurement import cell"
                )
            elif field == "quantity":
                normalized[field] = format_quantity(
                    parse_quantity(_spreadsheet_decimal_text(cell.value))
                )
            else:
                normalized[field] = _yuan_to_cents(cell.value)
        except (ApiError, InvalidOperation, TypeError, ValueError):
            errors.append(_cell_error(row_number, column, field, "单元格格式无效"))
    return normalized, errors


def _spreadsheet_positive_integer(value: object) -> int:
    if isinstance(value, bool):
        raise TypeError
    try:
        decimal = Decimal(str(value))
    except InvalidOperation:
        raise ValueError from None
    if not decimal.is_finite() or decimal != decimal.to_integral_value():
        raise ValueError
    return base._positive_integer(int(decimal), "Invalid procurement import cell")


def _spreadsheet_decimal_text(value: object) -> str:
    if isinstance(value, bool) or value is None:
        raise ValueError
    try:
        decimal = Decimal(str(value))
    except InvalidOperation:
        raise ValueError from None
    if not decimal.is_finite():
        raise ValueError
    return format(decimal, "f")


def _yuan_to_cents(value: object) -> int:
    decimal = Decimal(_spreadsheet_decimal_text(value))
    cents = decimal * 100
    if cents != cents.to_integral_value():
        raise ValueError
    return base._money(int(cents), "Invalid procurement import cell")


def _cell_error(row: int, column: int, field: str, message: str) -> dict[str, object]:
    return {"row": row, "column": column, "field": field, "message": message}


def _stored_preview(row: sqlite3.Row) -> dict[str, list[dict[str, object]]]:
    try:
        value = json.loads(row["preview_json"])
    except (TypeError, ValueError):
        raise sqlite3.DatabaseError("procurement import preview is invalid") from None
    if not isinstance(value, dict) or set(value) != {"rows", "errors"}:
        raise sqlite3.DatabaseError("procurement import preview is invalid")
    rows = value["rows"]
    errors = value["errors"]
    if not isinstance(rows, list) or not isinstance(errors, list):
        raise sqlite3.DatabaseError("procurement import preview is invalid")
    return {"rows": rows, "errors": errors}


def _import_row(
    connection: sqlite3.Connection, import_id: int, project_id: int
) -> sqlite3.Row | None:
    return connection.execute(
        "SELECT * FROM procurement_imports WHERE id = ? AND project_id = ?",
        (import_id, project_id),
    ).fetchone()


def _import_response(row: sqlite3.Row, project_code: str) -> dict[str, object]:
    preview = _stored_preview(row)
    return {
        "id": row["id"],
        "project_code": project_code,
        "filename": row["filename"],
        "sha256": row["sha256"],
        "status": row["status"],
        "revision": row["revision"],
        "expires_at": row["expires_at"],
        "confirmed_list_id": row["confirmed_list_id"],
        "rows": preview["rows"],
        "errors": preview["errors"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def _project_scope(method: str, project_code: str, suffix: str) -> str:
    return f"{method}:/api/projects/{project_code}{suffix}"


def _require_active_project(project: sqlite3.Row) -> None:
    if project["status"] != "active":
        raise base._business_conflict("Project is archived", "PROJECT_ARCHIVED")


def _derived_storage_key(storage_key: str, suffix: str) -> str:
    return hashlib.sha256(f"{storage_key}\0{suffix}".encode()).hexdigest()


def _inventory_exact_matches(
    connection: sqlite3.Connection,
) -> dict[tuple[str | None, str, str | None, str | None, str], int]:
    matches: dict[tuple[str | None, str, str | None, str | None, str], int] = {}
    ambiguous: set[tuple[str | None, str, str | None, str | None, str]] = set()
    for row in connection.execute(
        "SELECT id, brand, name, model, specification, unit FROM inventory_items"
    ).fetchall():
        identity = _inventory_identity(
            row["brand"],
            row["name"],
            row["model"],
            row["specification"],
            row["unit"],
        )
        if identity in matches:
            ambiguous.add(identity)
        else:
            matches[identity] = int(row["id"])
    for identity in ambiguous:
        matches.pop(identity, None)
    return matches


def _inventory_identity_from_line(
    line: dict[str, object],
) -> tuple[str | None, str, str | None, str | None, str]:
    return _inventory_identity(
        line["brand"],
        line["name"],
        line["model"],
        line["specification"],
        line["unit"],
    )


def _inventory_identity(
    brand: object,
    name: object,
    model: object,
    specification: object,
    unit: object,
) -> tuple[str | None, str, str | None, str | None, str]:
    normalized_name = _identity_text(name)
    normalized_unit = _identity_text(unit)
    if normalized_name is None or normalized_unit is None:
        raise sqlite3.DatabaseError("inventory identity is invalid")
    return (
        _identity_text(brand),
        normalized_name,
        _identity_text(model),
        _identity_text(specification),
        normalized_unit,
    )


def _identity_text(value: object) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str):
        raise sqlite3.DatabaseError("inventory identity is invalid")
    normalized = value.strip().casefold()
    return normalized or None


def _validate_order_references(
    connection: sqlite3.Connection,
    project: sqlite3.Row,
    normalized: dict[str, object],
) -> None:
    if connection.execute(
        "SELECT 1 FROM companies WHERE id = ?",
        (normalized["supplier_company_id"],),
    ).fetchone() is None:
        raise base._not_found("Company not found")
    base._validate_document_versions(
        connection,
        str(project["project_code"]),
        normalized["document_version_ids"],
    )
    for line in normalized["lines"]:
        procurement_line = base._confirmed_procurement_line(
            connection,
            int(line["procurement_line_id"]),
            int(project["id"]),
        )
        if (
            int(line["quantity_milli"]) > int(procurement_line["quantity_milli"])
            and line["overage_reason"] is None
        ):
            raise base._business_conflict(
                "Over-ordering requires a reason", "OVER_ORDER_REASON_REQUIRED"
            )


def _insert_order_lines(
    connection: sqlite3.Connection,
    order_id: int,
    lines: list[dict[str, object]],
    timestamp: str,
) -> None:
    for line in lines:
        connection.execute(
            """
            INSERT INTO purchase_order_lines
                (purchase_order_id, procurement_line_id, quantity_milli,
                 received_quantity_milli, unit_cost_cents, overage_reason, created_at)
            VALUES (?, ?, ?, 0, ?, ?, ?)
            """,
            (
                order_id,
                line["procurement_line_id"],
                line["quantity_milli"],
                line["unit_cost_cents"],
                line["overage_reason"],
                timestamp,
            ),
        )


def _order_has_active_facts(connection: sqlite3.Connection, order_id: int) -> bool:
    row = connection.execute(
        """
        SELECT
            EXISTS(SELECT 1 FROM goods_receipts
                   WHERE purchase_order_id = ? AND status = 'active')
            OR EXISTS(SELECT 1 FROM supplier_payments
                      WHERE purchase_order_id = ? AND status = 'active')
            OR EXISTS(SELECT 1 FROM supplier_invoices
                      WHERE purchase_order_id = ? AND status = 'active')
        """,
        (order_id, order_id, order_id),
    ).fetchone()
    return bool(row[0])


def _normalize_allocations(value: object, detail: str) -> list[dict[str, int]]:
    if not isinstance(value, list) or not value:
        raise base._invalid_payload(detail)
    allocations: list[dict[str, int]] = []
    seen: set[int] = set()
    for raw in value:
        if not isinstance(raw, dict) or set(raw) != _ALLOCATION_FIELDS:
            raise base._invalid_payload(detail)
        line_id = base._positive_integer(raw["purchase_order_line_id"], detail)
        amount = _positive_money(raw["amount_cents"], detail)
        if line_id in seen:
            raise base._invalid_payload(detail)
        seen.add(line_id)
        allocations.append(
            {"purchase_order_line_id": line_id, "amount_cents": amount}
        )
    return allocations


def _positive_money(value: object, detail: str) -> int:
    amount = base._money(value, detail)
    if amount == 0:
        raise base._invalid_payload(detail)
    return amount


def _normalize_payment(payload: dict[str, Any]) -> dict[str, object]:
    detail = "Invalid supplier payment payload"
    amount = _positive_money(payload["amount_cents"], detail)
    allocations = _normalize_allocations(payload["allocations"], detail)
    if sum(item["amount_cents"] for item in allocations) != amount:
        raise base._invalid_payload(detail)
    return {
        "paid_on": base._business_date(payload["paid_on"], detail),
        "amount_cents": amount,
        "payment_method": base._required_text(payload["payment_method"], detail),
        "reference_no": base._optional_text(payload["reference_no"], detail),
        "allocations": allocations,
        "notes": base._optional_text(payload["notes"], detail),
    }


def _normalize_invoice(payload: dict[str, Any]) -> dict[str, object]:
    detail = "Invalid supplier invoice payload"
    amount = _positive_money(payload["amount_cents"], detail)
    allocations = _normalize_allocations(payload["allocations"], detail)
    if sum(item["amount_cents"] for item in allocations) != amount:
        raise base._invalid_payload(detail)
    raw_documents = payload["document_version_ids"]
    if not isinstance(raw_documents, list):
        raise base._invalid_payload(detail)
    document_ids = [base._positive_integer(value, detail) for value in raw_documents]
    if len(set(document_ids)) != len(document_ids):
        raise base._invalid_payload(detail)
    return {
        "invoice_no": base._required_text(payload["invoice_no"], detail),
        "invoiced_on": base._business_date(payload["invoiced_on"], detail),
        "amount_cents": amount,
        "allocations": allocations,
        "document_version_ids": document_ids,
    }


def _invoice_attachment_error(field: str, message: str) -> ApiError:
    return ApiError(
        status.HTTP_422_UNPROCESSABLE_CONTENT,
        "Invalid supplier invoice payload",
        "VALIDATION_ERROR",
        field_errors={field: [message]},
    )


def _read_idempotency_key(request: Request) -> str:
    values = request.headers.getlist("Idempotency-Key")
    if len(values) != 1:
        raise base._invalid_payload("Invalid Idempotency-Key")
    return base._validate_idempotency_key(values[0])


def _attachment_too_large(max_size_bytes: int) -> ApiError:
    return ApiError(
        status.HTTP_413_CONTENT_TOO_LARGE,
        "Document file is too large",
        "DOCUMENT_FILE_TOO_LARGE",
        field_errors={
            "files": [f"must not exceed {max_size_bytes // (1024 * 1024)} MB"]
        },
    )


def _attachment_batch_too_large(max_size_bytes: int) -> ApiError:
    return ApiError(
        status.HTTP_413_CONTENT_TOO_LARGE,
        "Document attachment batch is too large",
        "DOCUMENT_BATCH_TOO_LARGE",
        field_errors={
            "files": [
                (
                    "must contain at most 20 files whose combined size does not "
                    f"exceed {max_size_bytes // (1024 * 1024)} MB"
                )
            ]
        },
    )


def _payable_order(
    connection: sqlite3.Connection, order_id: int, project_id: int
) -> sqlite3.Row:
    order = base._order_row(connection, order_id, project_id)
    if order is None:
        raise base._not_found("Purchase order not found")
    if order["status"] not in {"confirmed", "partially_received", "received"}:
        raise base._business_conflict(
            "Purchase order is not payable", "PURCHASE_ORDER_NOT_PAYABLE"
        )
    return order


def _validate_allocation_capacity(
    connection: sqlite3.Connection,
    order_id: int,
    allocations: list[dict[str, int]],
    *,
    fact: str,
) -> None:
    if fact == "payment":
        allocation_table = "supplier_payment_allocations"
        fact_table = "supplier_payments"
        fact_id_column = "supplier_payment_id"
        error_code = "PAYMENT_AMOUNT_EXCEEDED"
    elif fact == "invoice":
        allocation_table = "supplier_invoice_allocations"
        fact_table = "supplier_invoices"
        fact_id_column = "supplier_invoice_id"
        error_code = "INVOICE_AMOUNT_EXCEEDED"
    else:
        raise ValueError("unsupported allocation fact")
    for allocation in allocations:
        order_line = connection.execute(
            """
            SELECT * FROM purchase_order_lines
            WHERE id = ? AND purchase_order_id = ?
            """,
            (allocation["purchase_order_line_id"], order_id),
        ).fetchone()
        if order_line is None:
            raise base._invalid_payload(
                f"Invalid supplier {fact} payload"
            )
        already_allocated = connection.execute(
            f"""
            SELECT COALESCE(SUM(allocations.amount_cents), 0)
            FROM {allocation_table} AS allocations
            JOIN {fact_table} AS facts
              ON facts.id = allocations.{fact_id_column}
            WHERE allocations.purchase_order_line_id = ?
              AND facts.status = 'active'
            """,
            (allocation["purchase_order_line_id"],),
        ).fetchone()[0]
        line_amount = line_value(
            int(order_line["unit_cost_cents"]), int(order_line["quantity_milli"])
        )
        if int(already_allocated) + allocation["amount_cents"] > line_amount:
            raise base._business_conflict(
                f"Supplier {fact} exceeds the purchase order", error_code
            )


def _payment_row(
    connection: sqlite3.Connection, payment_id: int, project_id: int
) -> sqlite3.Row | None:
    return connection.execute(
        """
        SELECT payments.*
        FROM supplier_payments AS payments
        JOIN purchase_orders AS orders ON orders.id = payments.purchase_order_id
        WHERE payments.id = ? AND orders.project_id = ?
        """,
        (payment_id, project_id),
    ).fetchone()


def _invoice_row(
    connection: sqlite3.Connection, invoice_id: int, project_id: int
) -> sqlite3.Row | None:
    return connection.execute(
        """
        SELECT invoices.*
        FROM supplier_invoices AS invoices
        JOIN purchase_orders AS orders ON orders.id = invoices.purchase_order_id
        WHERE invoices.id = ? AND orders.project_id = ?
        """,
        (invoice_id, project_id),
    ).fetchone()


async def _reversal_request(
    request: Request, idempotency_key: str
) -> tuple[str, dict[str, object]]:
    key = base._validate_idempotency_key(idempotency_key)
    payload = await base._read_json(
        request, _CANCEL_FIELDS, "Invalid reversal payload"
    )
    return key, {
        "reason": base._required_text(payload["reason"], "Invalid reversal payload"),
        "expected_revision": base._positive_integer(
            payload["expected_revision"], "Invalid reversal payload"
        ),
    }


def _receipt_row(
    connection: sqlite3.Connection, receipt_id: int, project_id: int
) -> sqlite3.Row | None:
    return connection.execute(
        """
        SELECT receipts.*
        FROM goods_receipts AS receipts
        JOIN purchase_orders AS orders ON orders.id = receipts.purchase_order_id
        WHERE receipts.id = ? AND orders.project_id = ?
        """,
        (receipt_id, project_id),
    ).fetchone()


def _reverse_receipt_inventory(
    connection: sqlite3.Connection,
    receipt: sqlite3.Row,
    project_id: int,
    reason: str,
    timestamp: str,
) -> None:
    lines = connection.execute(
        """
        SELECT receipt_lines.*, order_lines.procurement_line_id
        FROM goods_receipt_lines AS receipt_lines
        JOIN purchase_order_lines AS order_lines
          ON order_lines.id = receipt_lines.purchase_order_line_id
        WHERE receipt_lines.goods_receipt_id = ? ORDER BY receipt_lines.id
        """,
        (receipt["id"],),
    ).fetchall()
    for line in lines:
        item = connection.execute(
            "SELECT * FROM inventory_items WHERE id = ?", (line["inventory_item_id"],)
        ).fetchone()
        if item is None:
            raise sqlite3.DatabaseError("receipt inventory item is missing")
        quantity_after = int(item["quantity_milli"]) - int(line["quantity_milli"])
        value_after = int(item["inventory_value_cents"]) - int(line["value_cents"])
        if quantity_after < 0 or value_after < 0 or (quantity_after == 0 and value_after != 0):
            raise base._business_conflict(
                "Receipt reversal would make inventory negative",
                "RECEIPT_REVERSAL_INSUFFICIENT_INVENTORY",
            )
        cursor = connection.execute(
            """
            INSERT INTO inventory_movements
                (inventory_item_id, project_id, procurement_line_id,
                 movement_type, quantity_delta_milli, value_delta_cents,
                 quantity_after_milli, value_after_cents, source_type,
                 source_id, occurred_on, reason, created_at)
            VALUES (?, ?, ?, 'reversal', ?, ?, ?, ?,
                    'goods_receipt_reversal', ?, ?, ?, ?)
            """,
            (
                item["id"],
                project_id,
                line["procurement_line_id"],
                -int(line["quantity_milli"]),
                -int(line["value_cents"]),
                quantity_after,
                value_after,
                receipt["id"],
                timestamp[:10],
                reason,
                timestamp,
            ),
        )
        if cursor.lastrowid is None:
            raise sqlite3.DatabaseError("receipt reversal movement is missing")
        connection.execute(
            """
            UPDATE inventory_items
            SET quantity_milli = ?, inventory_value_cents = ?,
                revision = revision + 1, updated_at = ?
            WHERE id = ? AND revision = ?
            """,
            (
                quantity_after,
                value_after,
                timestamp,
                item["id"],
                item["revision"],
            ),
        )
        updated_line = connection.execute(
            """
            UPDATE purchase_order_lines
            SET received_quantity_milli = received_quantity_milli - ?
            WHERE id = ? AND received_quantity_milli >= ?
            """,
            (line["quantity_milli"], line["purchase_order_line_id"], line["quantity_milli"]),
        )
        if updated_line.rowcount != 1:
            raise sqlite3.DatabaseError("purchase order receipt balance is invalid")


def _refresh_order_status_after_reversal(
    connection: sqlite3.Connection, order_id: int, timestamp: str
) -> None:
    totals = connection.execute(
        """
        SELECT COALESCE(SUM(quantity_milli), 0) AS ordered,
               COALESCE(SUM(received_quantity_milli), 0) AS received
        FROM purchase_order_lines WHERE purchase_order_id = ?
        """,
        (order_id,),
    ).fetchone()
    if totals["received"] == 0:
        next_status = "confirmed"
    elif totals["received"] == totals["ordered"]:
        next_status = "received"
    else:
        next_status = "partially_received"
    connection.execute(
        """
        UPDATE purchase_orders
        SET status = ?, revision = revision + 1, updated_at = ? WHERE id = ?
        """,
        (next_status, timestamp, order_id),
    )


def _quote_row(
    connection: sqlite3.Connection, export_id: int, project_id: int
) -> sqlite3.Row | None:
    return connection.execute(
        """
        SELECT exports.*, companies.name AS customer_company_name
        FROM quote_exports AS exports
        JOIN companies ON companies.id = exports.customer_company_id
        WHERE exports.id = ? AND exports.project_id = ?
        """,
        (export_id, project_id),
    ).fetchone()


def _quote_response(row: sqlite3.Row, project_code: str) -> dict[str, object]:
    return {
        "id": row["id"],
        "project_code": project_code,
        "procurement_list_id": row["procurement_list_id"],
        "title": row["title"],
        "customer_company_id": row["customer_company_id"],
        "customer_company_name": row["customer_company_name"],
        "notes": row["notes"],
        "created_at": row["created_at"],
        "download_url": (
            f"/api/projects/{project_code}/quote-exports/{row['id']}/download"
        ),
    }


def _quote_workbook(connection: sqlite3.Connection, export: sqlite3.Row) -> bytes:
    lines = connection.execute(
        """
        SELECT sequence_no, category, name, specification, brand, model,
               quantity_milli, unit, quoted_unit_price_cents
        FROM procurement_lines
        WHERE procurement_list_id = ? ORDER BY sequence_no, id
        """,
        (export["procurement_list_id"],),
    ).fetchall()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "报价单"
    worksheet.append(("标题", export["title"]))
    worksheet.append(("客户公司", export["customer_company_name"]))
    worksheet.append(("备注", export["notes"]))
    worksheet.append((None,))
    worksheet.append(_CUSTOMER_QUOTE_HEADERS)
    for line in lines:
        quoted_total = line_value(
            int(line["quoted_unit_price_cents"]), int(line["quantity_milli"])
        )
        worksheet.append(
            (
                line["sequence_no"],
                line["category"],
                line["name"],
                line["specification"],
                line["brand"],
                line["model"],
                format_quantity(int(line["quantity_milli"])),
                line["unit"],
                _cents_text(int(line["quoted_unit_price_cents"])),
                _cents_text(quoted_total),
            )
        )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _cents_text(value: int) -> str:
    return f"{Decimal(value) / Decimal(100):.2f}"


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)
