from __future__ import annotations

import hashlib
import json
import sqlite3
from collections import Counter
from collections.abc import Callable
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any
from uuid import UUID

from fastapi import APIRouter, Depends, Header, Request, Response, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from backend.app.core.config import Settings
from backend.app.core.database import transaction_immediate
from backend.app.core.storage_paths import normalize_project_code, project_code_identity
from backend.app.features.api_common import (
    ApiError,
    ApiErrorRoute,
    idempotency_scope,
    idempotency_storage_key,
    restore_idempotent_response,
    save_idempotent_response,
)
from backend.app.features.auth import require_authenticated_session
from backend.app.features.inventory import (
    ensure_receipt_inventory_item,
    format_quantity,
    line_value,
    parse_quantity,
    post_receipt_movement,
)

Clock = Callable[[], datetime]

_LIST_FIELDS = {"name", "notes"}
_LINE_FIELDS = {
    "sequence_no",
    "category",
    "name",
    "specification",
    "brand",
    "model",
    "quantity",
    "unit",
    "unit_cost_cents",
    "quoted_unit_price_cents",
}
_LINE_UPDATE_FIELDS = {*_LINE_FIELDS, "expected_revision"}
_CONFIRM_FIELDS = {"expected_revision"}
_ORDER_FIELDS = {
    "order_no",
    "supplier_company_id",
    "ordered_on",
    "expected_delivery_on",
    "lines",
    "notes",
    "document_version_ids",
}
_ORDER_LINE_FIELDS = {
    "procurement_line_id",
    "quantity",
    "unit_cost_cents",
    "overage_reason",
}
_RECEIPT_FIELDS = {"received_on", "warehouse_name", "lines", "notes"}
_RECEIPT_LINE_FIELDS = {"purchase_order_line_id", "quantity"}
_TEMPLATE_HEADERS = (
    "序号",
    "大类",
    "名称",
    "规格",
    "品牌",
    "型号",
    "数量",
    "单位",
    "成本单价（元）",
    "报价单价（元）",
)
_SQLITE_MAX_INTEGER = 2**63 - 1
_MAX_PAGE_SIZE = 200


def create_procurement_router(
    get_connection: Callable[..., sqlite3.Connection],
    get_settings: Callable[..., Settings],
    *,
    clock: Clock | None = None,
) -> APIRouter:
    router = APIRouter(route_class=ApiErrorRoute, tags=["procurement"])
    connection_dependency = Depends(get_connection)
    settings_dependency = Depends(get_settings)
    now = clock or _utc_now

    def require_session(
        request: Request,
        settings: Settings = settings_dependency,
    ) -> None:
        require_authenticated_session(request, settings.session_secret)

    authentication_dependency = Depends(require_session)

    @router.get("/api/procurement/import-template.xlsx")
    def download_procurement_template(
        _: None = authentication_dependency,
    ) -> StreamingResponse:
        content = _template_workbook()
        return StreamingResponse(
            BytesIO(content),
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": (
                    "attachment; filename=procurement-import-template.xlsx"
                )
            },
        )

    @router.get("/api/projects/{project_code}/procurement-lists")
    def list_procurement_lists(
        project_code: str,
        request: Request,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        project = _project(connection, _normalize_project_path(project_code))
        page, page_size = _read_pagination(request)
        total = connection.execute(
            "SELECT COUNT(*) FROM procurement_lists WHERE project_id = ?",
            (project["id"],),
        ).fetchone()[0]
        rows = connection.execute(
            """
            SELECT lists.*, COUNT(lines.id) AS line_count
            FROM procurement_lists AS lists
            LEFT JOIN procurement_lines AS lines
                ON lines.procurement_list_id = lists.id
            WHERE lists.project_id = ?
            GROUP BY lists.id
            ORDER BY lists.created_at DESC, lists.id DESC
            LIMIT ? OFFSET ?
            """,
            (project["id"], page_size, (page - 1) * page_size),
        ).fetchall()
        return {
            "items": [_list_summary(row, project_code) for row in rows],
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    @router.post(
        "/api/projects/{project_code}/procurement-lists",
        status_code=status.HTTP_201_CREATED,
    )
    async def create_procurement_list(
        project_code: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        key = _validate_idempotency_key(idempotency_key)
        payload = await _read_json(request, _LIST_FIELDS, "Invalid procurement payload")
        normalized = {
            "name": _required_text(payload["name"], "Invalid procurement payload"),
            "notes": _optional_text(payload["notes"], "Invalid procurement payload"),
        }
        request_hash = _request_hash(normalized)
        scope = idempotency_scope(request)
        storage_key = idempotency_storage_key(scope, key)
        project_key = _normalize_project_path(project_code)
        timestamp = _timestamp(now)
        with transaction_immediate(connection):
            restored = restore_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
            )
            if restored is not None:
                return restored
            project = _active_project(connection, project_key)
            cursor = connection.execute(
                """
                INSERT INTO procurement_lists
                    (project_id, name, notes, status, revision,
                     create_idempotency_key, create_request_hash,
                     created_at, updated_at)
                VALUES (?, ?, ?, 'draft', 1, ?, ?, ?, ?)
                """,
                (
                    project["id"],
                    normalized["name"],
                    normalized["notes"],
                    storage_key,
                    request_hash,
                    timestamp,
                    timestamp,
                ),
            )
            row = _list_row(connection, _last_insert_id(cursor), int(project["id"]))
            if row is None:
                raise sqlite3.DatabaseError("created procurement list is missing")
            response = _list_detail(connection, row, str(project["project_code"]))
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_201_CREATED,
                resource_type="procurement_list",
                resource_id=int(row["id"]),
                created_at=timestamp,
            )
            return response

    @router.get("/api/projects/{project_code}/procurement-lists/{list_id}")
    def get_procurement_list(
        project_code: str,
        list_id: str,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        project = _project(connection, _normalize_project_path(project_code))
        row = _list_row(connection, _parse_identifier(list_id), int(project["id"]))
        if row is None:
            raise _not_found("Procurement list not found")
        return _list_detail(connection, row, str(project["project_code"]))

    @router.put("/api/projects/{project_code}/procurement-lists/{list_id}")
    async def update_procurement_list(
        project_code: str,
        list_id: str,
        request: Request,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = _parse_identifier(list_id)
        payload = await _read_json(
            request,
            {*_LIST_FIELDS, "expected_revision"},
            "Invalid procurement payload",
        )
        normalized = {
            "name": _required_text(payload["name"], "Invalid procurement payload"),
            "notes": _optional_text(payload["notes"], "Invalid procurement payload"),
        }
        expected_revision = _positive_integer(
            payload["expected_revision"], "Invalid procurement payload"
        )
        timestamp = _timestamp(now)
        with transaction_immediate(connection):
            project = _active_project(connection, _normalize_project_path(project_code))
            row = _editable_list(connection, identifier, int(project["id"]))
            _require_revision(row, expected_revision)
            connection.execute(
                """
                UPDATE procurement_lists
                SET name = ?, notes = ?, revision = revision + 1, updated_at = ?
                WHERE id = ? AND revision = ?
                """,
                (
                    normalized["name"],
                    normalized["notes"],
                    timestamp,
                    identifier,
                    expected_revision,
                ),
            )
            updated = _list_row(connection, identifier, int(project["id"]))
            return _list_detail(connection, updated, str(project["project_code"]))

    @router.post(
        "/api/projects/{project_code}/procurement-lists/{list_id}/lines",
        status_code=status.HTTP_201_CREATED,
    )
    async def create_procurement_line(
        project_code: str,
        list_id: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = _parse_identifier(list_id)
        key = _validate_idempotency_key(idempotency_key)
        payload = await _read_json(request, _LINE_FIELDS, "Invalid procurement payload")
        normalized = _normalize_line(payload)
        request_hash = _request_hash(normalized)
        scope = idempotency_scope(request)
        storage_key = idempotency_storage_key(scope, key)
        timestamp = _timestamp(now)
        try:
            with transaction_immediate(connection):
                restored = restore_idempotent_response(
                    connection,
                    scope=scope,
                    key=key,
                    request_hash=request_hash,
                )
                if restored is not None:
                    return restored
                project = _active_project(
                    connection, _normalize_project_path(project_code)
                )
                _editable_list(
                    connection,
                    identifier,
                    int(project["id"]),
                )
                cursor = connection.execute(
                    """
                    INSERT INTO procurement_lines
                        (procurement_list_id, inventory_item_id, sequence_no,
                         category, name, specification, brand, model,
                         quantity_milli, unit, unit_cost_cents,
                         quoted_unit_price_cents, revision,
                         create_idempotency_key, create_request_hash,
                         created_at, updated_at)
                    VALUES (?, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?)
                    """,
                    (
                        identifier,
                        normalized["sequence_no"],
                        normalized["category"],
                        normalized["name"],
                        normalized["specification"],
                        normalized["brand"],
                        normalized["model"],
                        normalized["quantity_milli"],
                        normalized["unit"],
                        normalized["unit_cost_cents"],
                        normalized["quoted_unit_price_cents"],
                        storage_key,
                        request_hash,
                        timestamp,
                        timestamp,
                    ),
                )
                line_id = _last_insert_id(cursor)
                connection.execute(
                    """
                    UPDATE procurement_lists
                    SET revision = revision + 1, updated_at = ? WHERE id = ?
                    """,
                    (timestamp, identifier),
                )
                row = _line_row(connection, line_id, identifier)
                updated_list = _list_row(connection, identifier, int(project["id"]))
                response = _line_response(
                    connection, row, int(updated_list["revision"])
                )
                save_idempotent_response(
                    connection,
                    scope=scope,
                    key=key,
                    request_hash=request_hash,
                    response=response,
                    response_status=status.HTTP_201_CREATED,
                    resource_type="procurement_line",
                    resource_id=line_id,
                    created_at=timestamp,
                )
                return response
        except sqlite3.IntegrityError as exc:
            if _is_unique_constraint(exc):
                raise _business_conflict(
                    "Procurement sequence already exists",
                    "PROCUREMENT_SEQUENCE_EXISTS",
                ) from None
            raise

    @router.put(
        "/api/projects/{project_code}/procurement-lists/{list_id}/lines/{line_id}"
    )
    async def update_procurement_line(
        project_code: str,
        list_id: str,
        line_id: str,
        request: Request,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        list_identifier = _parse_identifier(list_id)
        line_identifier = _parse_identifier(line_id)
        payload = await _read_json(
            request,
            _LINE_UPDATE_FIELDS,
            "Invalid procurement payload",
        )
        expected_revision = _positive_integer(
            payload.pop("expected_revision"), "Invalid procurement payload"
        )
        normalized = _normalize_line(payload)
        timestamp = _timestamp(now)
        try:
            with transaction_immediate(connection):
                project = _active_project(
                    connection, _normalize_project_path(project_code)
                )
                _editable_list(connection, list_identifier, int(project["id"]))
                row = _line_row(connection, line_identifier, list_identifier)
                if row is None:
                    raise _not_found("Procurement line not found")
                _require_revision(row, expected_revision)
                connection.execute(
                    """
                    UPDATE procurement_lines
                    SET sequence_no = ?, category = ?, name = ?,
                        specification = ?, brand = ?, model = ?,
                        quantity_milli = ?, unit = ?, unit_cost_cents = ?,
                        quoted_unit_price_cents = ?, revision = revision + 1,
                        updated_at = ?
                    WHERE id = ? AND revision = ?
                    """,
                    (
                        normalized["sequence_no"],
                        normalized["category"],
                        normalized["name"],
                        normalized["specification"],
                        normalized["brand"],
                        normalized["model"],
                        normalized["quantity_milli"],
                        normalized["unit"],
                        normalized["unit_cost_cents"],
                        normalized["quoted_unit_price_cents"],
                        timestamp,
                        line_identifier,
                        expected_revision,
                    ),
                )
                connection.execute(
                    """
                    UPDATE procurement_lists
                    SET revision = revision + 1, updated_at = ? WHERE id = ?
                    """,
                    (timestamp, list_identifier),
                )
                updated = _line_row(connection, line_identifier, list_identifier)
                updated_list = _list_row(
                    connection,
                    list_identifier,
                    int(project["id"]),
                )
                return _line_response(
                    connection, updated, int(updated_list["revision"])
                )
        except sqlite3.IntegrityError as exc:
            if _is_unique_constraint(exc):
                raise _business_conflict(
                    "Procurement sequence already exists",
                    "PROCUREMENT_SEQUENCE_EXISTS",
                ) from None
            raise

    @router.delete(
        "/api/projects/{project_code}/procurement-lists/{list_id}/lines/{line_id}",
        status_code=status.HTTP_204_NO_CONTENT,
    )
    def delete_procurement_line(
        project_code: str,
        list_id: str,
        line_id: str,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> Response:
        list_identifier = _parse_identifier(list_id)
        line_identifier = _parse_identifier(line_id)
        timestamp = _timestamp(now)
        with transaction_immediate(connection):
            project = _active_project(connection, _normalize_project_path(project_code))
            _editable_list(connection, list_identifier, int(project["id"]))
            cursor = connection.execute(
                """
                DELETE FROM procurement_lines
                WHERE id = ? AND procurement_list_id = ?
                """,
                (line_identifier, list_identifier),
            )
            if cursor.rowcount != 1:
                raise _not_found("Procurement line not found")
            connection.execute(
                """
                UPDATE procurement_lists
                SET revision = revision + 1, updated_at = ? WHERE id = ?
                """,
                (timestamp, list_identifier),
            )
        return Response(status_code=status.HTTP_204_NO_CONTENT)

    @router.post("/api/projects/{project_code}/procurement-lists/{list_id}/confirm")
    async def confirm_procurement_list(
        project_code: str,
        list_id: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = _parse_identifier(list_id)
        key = _validate_idempotency_key(idempotency_key)
        payload = await _read_json(
            request, _CONFIRM_FIELDS, "Invalid procurement payload"
        )
        expected_revision = _positive_integer(
            payload["expected_revision"], "Invalid procurement payload"
        )
        request_hash = _request_hash({"expected_revision": expected_revision})
        scope = idempotency_scope(request)
        storage_key = idempotency_storage_key(scope, key)
        timestamp = _timestamp(now)
        with transaction_immediate(connection):
            restored = restore_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
            )
            if restored is not None:
                return restored
            project = _active_project(connection, _normalize_project_path(project_code))
            row = _list_row(connection, identifier, int(project["id"]))
            if row is None:
                raise _not_found("Procurement list not found")
            if row["confirm_idempotency_key"] == key:
                _require_same_hash(row["confirm_request_hash"], request_hash)
                return _list_detail(connection, row, str(project["project_code"]))
            if row["status"] != "draft":
                raise _business_conflict(
                    "Procurement list is not editable",
                    "PROCUREMENT_LIST_NOT_EDITABLE",
                )
            _require_revision(row, expected_revision)
            if (
                connection.execute(
                    "SELECT 1 FROM procurement_lines WHERE procurement_list_id = ? LIMIT 1",
                    (identifier,),
                ).fetchone()
                is None
            ):
                raise _business_conflict(
                    "Procurement list has no lines",
                    "PROCUREMENT_LIST_EMPTY",
                )
            connection.execute(
                """
                UPDATE procurement_lists
                SET status = 'confirmed', confirmed_at = ?,
                    confirm_idempotency_key = ?, confirm_request_hash = ?,
                    revision = revision + 1, updated_at = ?
                WHERE id = ?
                """,
                (timestamp, storage_key, request_hash, timestamp, identifier),
            )
            confirmed = _list_row(connection, identifier, int(project["id"]))
            response = _list_detail(connection, confirmed, str(project["project_code"]))
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_200_OK,
                resource_type="procurement_list",
                resource_id=identifier,
                created_at=timestamp,
            )
            return response

    @router.get("/api/projects/{project_code}/purchase-orders")
    def list_purchase_orders(
        project_code: str,
        request: Request,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        project = _project(connection, _normalize_project_path(project_code))
        page, page_size = _read_pagination(request)
        selected_status = _single_query(request, "status")
        if selected_status is not None and selected_status not in {
            "draft",
            "confirmed",
            "partially_received",
            "received",
            "cancelled",
        }:
            raise _invalid_payload("Invalid purchase order status")
        where = "project_id = ?"
        parameters: list[object] = [project["id"]]
        if selected_status is not None:
            where += " AND status = ?"
            parameters.append(selected_status)
        total = connection.execute(
            f"SELECT COUNT(*) FROM purchase_orders WHERE {where}",
            parameters,
        ).fetchone()[0]
        rows = connection.execute(
            f"""
            SELECT * FROM purchase_orders WHERE {where}
            ORDER BY ordered_on DESC, id DESC LIMIT ? OFFSET ?
            """,
            (*parameters, page_size, (page - 1) * page_size),
        ).fetchall()
        return {
            "items": [
                _order_response(connection, row, str(project["project_code"]))
                for row in rows
            ],
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    @router.post(
        "/api/projects/{project_code}/purchase-orders",
        status_code=status.HTTP_201_CREATED,
    )
    async def create_purchase_order(
        project_code: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        key = _validate_idempotency_key(idempotency_key)
        payload = await _read_json(
            request, _ORDER_FIELDS, "Invalid purchase order payload"
        )
        normalized = _normalize_order(payload)
        request_hash = _request_hash(normalized)
        scope = idempotency_scope(request)
        storage_key = idempotency_storage_key(scope, key)
        timestamp = _timestamp(now)
        try:
            with transaction_immediate(connection):
                restored = restore_idempotent_response(
                    connection,
                    scope=scope,
                    key=key,
                    request_hash=request_hash,
                )
                if restored is not None:
                    return restored
                project = _active_project(
                    connection, _normalize_project_path(project_code)
                )
                if (
                    connection.execute(
                        "SELECT 1 FROM companies WHERE id = ?",
                        (normalized["supplier_company_id"],),
                    ).fetchone()
                    is None
                ):
                    raise _not_found("Company not found")
                _validate_document_versions(
                    connection,
                    str(project["project_code"]),
                    normalized["document_version_ids"],
                )
                cursor = connection.execute(
                    """
                    INSERT INTO purchase_orders
                        (project_id, order_no, supplier_company_id, ordered_on,
                         expected_delivery_on, notes, status, revision,
                         create_idempotency_key, create_request_hash,
                         created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, 'draft', 1, ?, ?, ?, ?)
                    """,
                    (
                        project["id"],
                        normalized["order_no"],
                        normalized["supplier_company_id"],
                        normalized["ordered_on"],
                        normalized["expected_delivery_on"],
                        normalized["notes"],
                        storage_key,
                        request_hash,
                        timestamp,
                        timestamp,
                    ),
                )
                order_id = _last_insert_id(cursor)
                for line in normalized["lines"]:
                    procurement_line = _confirmed_procurement_line(
                        connection,
                        int(line["procurement_line_id"]),
                        int(project["id"]),
                    )
                    if (
                        int(line["quantity_milli"])
                        > int(procurement_line["quantity_milli"])
                        and line["overage_reason"] is None
                    ):
                        raise _business_conflict(
                            "Over-ordering requires a reason",
                            "OVER_ORDER_REASON_REQUIRED",
                        )
                    connection.execute(
                        """
                        INSERT INTO purchase_order_lines
                            (purchase_order_id, procurement_line_id,
                             quantity_milli, received_quantity_milli,
                             unit_cost_cents, overage_reason, created_at)
                        VALUES (?, ?, ?, 0, ?, ?, ?)
                        """,
                        (
                            order_id,
                            line["procurement_line_id"],
                            line["quantity_milli"],
                            line["unit_cost_cents"],
                            line["overage_reason"],
                            timestamp,
                        ),
                    )
                for document_id in normalized["document_version_ids"]:
                    connection.execute(
                        """
                        INSERT INTO purchase_order_documents
                            (purchase_order_id, document_version_id)
                        VALUES (?, ?)
                        """,
                        (order_id, document_id),
                    )
                order = _order_row(connection, order_id, int(project["id"]))
                if order is None:
                    raise sqlite3.DatabaseError("created purchase order is missing")
                response = _order_response(
                    connection, order, str(project["project_code"])
                )
                save_idempotent_response(
                    connection,
                    scope=scope,
                    key=key,
                    request_hash=request_hash,
                    response=response,
                    response_status=status.HTTP_201_CREATED,
                    resource_type="purchase_order",
                    resource_id=order_id,
                    created_at=timestamp,
                )
                return response
        except sqlite3.IntegrityError as exc:
            if _is_unique_constraint(exc):
                raise _business_conflict(
                    "Purchase order already exists",
                    "PURCHASE_ORDER_EXISTS",
                ) from None
            raise

    @router.get("/api/projects/{project_code}/purchase-orders/{order_id}")
    def get_purchase_order(
        project_code: str,
        order_id: str,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        project = _project(connection, _normalize_project_path(project_code))
        order = _order_row(connection, _parse_identifier(order_id), int(project["id"]))
        if order is None:
            raise _not_found("Purchase order not found")
        return _order_response(connection, order, str(project["project_code"]))

    @router.post("/api/projects/{project_code}/purchase-orders/{order_id}/confirm")
    async def confirm_purchase_order(
        project_code: str,
        order_id: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = _parse_identifier(order_id)
        key = _validate_idempotency_key(idempotency_key)
        payload = await _read_json(
            request, _CONFIRM_FIELDS, "Invalid purchase order payload"
        )
        expected_revision = _positive_integer(
            payload["expected_revision"], "Invalid purchase order payload"
        )
        request_hash = _request_hash({"expected_revision": expected_revision})
        scope = idempotency_scope(request)
        storage_key = idempotency_storage_key(scope, key)
        timestamp = _timestamp(now)
        with transaction_immediate(connection):
            restored = restore_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
            )
            if restored is not None:
                return restored
            project = _active_project(connection, _normalize_project_path(project_code))
            order = _order_row(connection, identifier, int(project["id"]))
            if order is None:
                raise _not_found("Purchase order not found")
            if order["confirm_idempotency_key"] == key:
                _require_same_hash(order["confirm_request_hash"], request_hash)
                return _order_response(connection, order, str(project["project_code"]))
            if order["status"] != "draft":
                raise _business_conflict(
                    "Purchase order cannot be confirmed",
                    "PURCHASE_ORDER_NOT_CONFIRMABLE",
                )
            _require_revision(order, expected_revision)
            for line in _order_lines(connection, identifier):
                previously_ordered = connection.execute(
                    """
                    SELECT COALESCE(SUM(order_lines.quantity_milli), 0)
                    FROM purchase_order_lines AS order_lines
                    JOIN purchase_orders AS orders
                        ON orders.id = order_lines.purchase_order_id
                    WHERE order_lines.procurement_line_id = ?
                      AND orders.status IN (
                          'confirmed', 'partially_received', 'received'
                      )
                    """,
                    (line["procurement_line_id"],),
                ).fetchone()[0]
                requirement = connection.execute(
                    "SELECT quantity_milli FROM procurement_lines WHERE id = ?",
                    (line["procurement_line_id"],),
                ).fetchone()[0]
                if (
                    previously_ordered + line["quantity_milli"] > requirement
                    and line["overage_reason"] is None
                ):
                    raise _business_conflict(
                        "Over-ordering requires a reason",
                        "OVER_ORDER_REASON_REQUIRED",
                    )
            connection.execute(
                """
                UPDATE purchase_orders
                SET status = 'confirmed', confirmed_at = ?,
                    confirm_idempotency_key = ?, confirm_request_hash = ?,
                    revision = revision + 1, updated_at = ?
                WHERE id = ?
                """,
                (timestamp, storage_key, request_hash, timestamp, identifier),
            )
            confirmed = _order_row(connection, identifier, int(project["id"]))
            response = _order_response(
                connection, confirmed, str(project["project_code"])
            )
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_200_OK,
                resource_type="purchase_order",
                resource_id=identifier,
                created_at=timestamp,
            )
            return response

    @router.post(
        "/api/projects/{project_code}/purchase-orders/{order_id}/goods-receipts",
        status_code=status.HTTP_201_CREATED,
    )
    async def create_goods_receipt(
        project_code: str,
        order_id: str,
        request: Request,
        idempotency_key: str = Header(alias="Idempotency-Key"),
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        identifier = _parse_identifier(order_id)
        key = _validate_idempotency_key(idempotency_key)
        payload = await _read_json(
            request, _RECEIPT_FIELDS, "Invalid goods receipt payload"
        )
        normalized = _normalize_receipt(payload)
        request_hash = _request_hash(normalized)
        scope = idempotency_scope(request)
        storage_key = idempotency_storage_key(scope, key)
        timestamp = _timestamp(now)
        with transaction_immediate(connection):
            restored = restore_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
            )
            if restored is not None:
                return restored
            project = _active_project(connection, _normalize_project_path(project_code))
            order = _order_row(connection, identifier, int(project["id"]))
            if order is None:
                raise _not_found("Purchase order not found")
            if order["status"] not in {
                "confirmed",
                "partially_received",
                "received",
            }:
                raise _business_conflict(
                    "Purchase order cannot receive goods",
                    "PURCHASE_ORDER_NOT_RECEIVABLE",
                )
            cursor = connection.execute(
                """
                INSERT INTO goods_receipts
                    (purchase_order_id, received_on, warehouse_name, notes,
                     status, revision, idempotency_key, request_hash,
                     created_at, updated_at)
                VALUES (?, ?, ?, ?, 'active', 1, ?, ?, ?, ?)
                """,
                (
                    identifier,
                    normalized["received_on"],
                    normalized["warehouse_name"],
                    normalized["notes"],
                    storage_key,
                    request_hash,
                    timestamp,
                    timestamp,
                ),
            )
            receipt_id = _last_insert_id(cursor)
            for input_line in normalized["lines"]:
                order_line = connection.execute(
                    """
                    SELECT order_lines.*, procurement_lines.inventory_item_id,
                           procurement_lines.name, procurement_lines.unit,
                           procurement_lines.brand, procurement_lines.model,
                           procurement_lines.specification
                    FROM purchase_order_lines AS order_lines
                    JOIN procurement_lines
                        ON procurement_lines.id = order_lines.procurement_line_id
                    WHERE order_lines.id = ?
                      AND order_lines.purchase_order_id = ?
                    """,
                    (input_line["purchase_order_line_id"], identifier),
                ).fetchone()
                if order_line is None:
                    raise _invalid_payload("Invalid goods receipt payload")
                quantity = int(input_line["quantity_milli"])
                if (
                    order_line["received_quantity_milli"] + quantity
                    > order_line["quantity_milli"]
                ):
                    raise _business_conflict(
                        "Receipt quantity exceeds ordered quantity",
                        "RECEIPT_QUANTITY_EXCEEDED",
                    )
                item = ensure_receipt_inventory_item(
                    connection,
                    order_line,
                    timestamp=timestamp,
                )
                value_cents = line_value(int(order_line["unit_cost_cents"]), quantity)
                movement_id = post_receipt_movement(
                    connection,
                    item=item,
                    project_id=int(project["id"]),
                    procurement_line_id=int(order_line["procurement_line_id"]),
                    receipt_id=receipt_id,
                    quantity_milli=quantity,
                    value_cents=value_cents,
                    received_on=str(normalized["received_on"]),
                    created_at=timestamp,
                )
                connection.execute(
                    """
                    INSERT INTO goods_receipt_lines
                        (goods_receipt_id, purchase_order_line_id,
                         inventory_item_id, quantity_milli, value_cents,
                         movement_id)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        receipt_id,
                        order_line["id"],
                        item["id"],
                        quantity,
                        value_cents,
                        movement_id,
                    ),
                )
                connection.execute(
                    """
                    UPDATE purchase_order_lines
                    SET received_quantity_milli = received_quantity_milli + ?
                    WHERE id = ?
                    """,
                    (quantity, order_line["id"]),
                )
            _refresh_order_receipt_status(connection, identifier, timestamp)
            receipt = connection.execute(
                "SELECT * FROM goods_receipts WHERE id = ?",
                (receipt_id,),
            ).fetchone()
            if receipt is None:
                raise sqlite3.DatabaseError("created goods receipt is missing")
            response = _receipt_response(connection, receipt)
            save_idempotent_response(
                connection,
                scope=scope,
                key=key,
                request_hash=request_hash,
                response=response,
                response_status=status.HTTP_201_CREATED,
                resource_type="goods_receipt",
                resource_id=receipt_id,
                created_at=timestamp,
            )
            return response

    @router.get("/api/projects/{project_code}/procurement-overview")
    def get_procurement_overview(
        project_code: str,
        _: None = authentication_dependency,
        connection: sqlite3.Connection = connection_dependency,
    ) -> dict[str, object]:
        project = _project(connection, _normalize_project_path(project_code))
        line_rows = _project_line_rows(connection, int(project["id"]))
        line_responses = [_line_response(connection, row, None) for row in line_rows]
        counts = Counter(str(line["order_status"]) for line in line_responses)
        for label in ("not_ordered", "partial", "ordered", "over_ordered"):
            counts.setdefault(label, 0)
        committed = connection.execute(
            """
            SELECT COALESCE(SUM(
                (order_lines.unit_cost_cents * order_lines.quantity_milli + 500) / 1000
            ), 0)
            FROM purchase_order_lines AS order_lines
            JOIN purchase_orders AS orders ON orders.id = order_lines.purchase_order_id
            WHERE orders.project_id = ?
              AND orders.status IN ('confirmed', 'partially_received', 'received')
            """,
            (project["id"],),
        ).fetchone()[0]
        received = connection.execute(
            """
            SELECT COALESCE(SUM(receipt_lines.value_cents), 0)
            FROM goods_receipt_lines AS receipt_lines
            JOIN goods_receipts AS receipts
                ON receipts.id = receipt_lines.goods_receipt_id
            JOIN purchase_orders AS orders
                ON orders.id = receipts.purchase_order_id
            WHERE orders.project_id = ? AND receipts.status = 'active'
            """,
            (project["id"],),
        ).fetchone()[0]
        consumed = connection.execute(
            """
            SELECT COALESCE(SUM(total_cost_cents), 0)
            FROM inventory_issues
            WHERE project_id = ? AND status = 'active'
            """,
            (project["id"],),
        ).fetchone()[0]
        return {
            "project_code": project["project_code"],
            "line_count": len(line_responses),
            "line_status_counts": dict(counts),
            "procurement_committed_cents": committed,
            "procurement_received_cents": received,
            "procurement_paid_cents": 0,
            "material_consumed_cents": consumed,
        }

    return router


def _template_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "采购清单"
    worksheet.append(_TEMPLATE_HEADERS)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:J1"
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for column, width in {
        "A": 8,
        "B": 14,
        "C": 20,
        "D": 18,
        "E": 14,
        "F": 18,
        "G": 12,
        "H": 10,
        "I": 18,
        "J": 18,
    }.items():
        worksheet.column_dimensions[column].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _normalize_line(payload: dict[str, Any]) -> dict[str, object]:
    quantity_milli = parse_quantity(payload["quantity"])
    unit_cost_cents = _money(payload["unit_cost_cents"], "Invalid procurement payload")
    quoted_unit_price_cents = _money(
        payload["quoted_unit_price_cents"], "Invalid procurement payload"
    )
    line_value(
        unit_cost_cents,
        quantity_milli,
        detail="Invalid procurement payload",
    )
    line_value(
        quoted_unit_price_cents,
        quantity_milli,
        detail="Invalid procurement payload",
    )
    return {
        "sequence_no": _positive_integer(
            payload["sequence_no"], "Invalid procurement payload"
        ),
        "category": _required_text(payload["category"], "Invalid procurement payload"),
        "name": _required_text(payload["name"], "Invalid procurement payload"),
        "specification": _optional_text(
            payload["specification"], "Invalid procurement payload"
        ),
        "brand": _optional_text(payload["brand"], "Invalid procurement payload"),
        "model": _optional_text(payload["model"], "Invalid procurement payload"),
        "quantity_milli": quantity_milli,
        "unit": _required_text(payload["unit"], "Invalid procurement payload"),
        "unit_cost_cents": unit_cost_cents,
        "quoted_unit_price_cents": quoted_unit_price_cents,
    }


def _normalize_order(payload: dict[str, Any]) -> dict[str, object]:
    raw_lines = payload["lines"]
    if not isinstance(raw_lines, list) or not raw_lines:
        raise _invalid_payload("Invalid purchase order payload")
    lines: list[dict[str, object]] = []
    seen_lines: set[int] = set()
    for raw_line in raw_lines:
        if not isinstance(raw_line, dict) or set(raw_line) != _ORDER_LINE_FIELDS:
            raise _invalid_payload("Invalid purchase order payload")
        procurement_line_id = _positive_integer(
            raw_line["procurement_line_id"], "Invalid purchase order payload"
        )
        if procurement_line_id in seen_lines:
            raise _invalid_payload("Invalid purchase order payload")
        seen_lines.add(procurement_line_id)
        quantity_milli = parse_quantity(raw_line["quantity"])
        unit_cost_cents = _money(
            raw_line["unit_cost_cents"], "Invalid purchase order payload"
        )
        line_value(
            unit_cost_cents,
            quantity_milli,
            detail="Invalid purchase order payload",
        )
        lines.append(
            {
                "procurement_line_id": procurement_line_id,
                "quantity_milli": quantity_milli,
                "unit_cost_cents": unit_cost_cents,
                "overage_reason": _optional_text(
                    raw_line["overage_reason"], "Invalid purchase order payload"
                ),
            }
        )
    order_total = sum(
        line_value(
            int(line["unit_cost_cents"]),
            int(line["quantity_milli"]),
            detail="Invalid purchase order payload",
        )
        for line in lines
    )
    if order_total > _SQLITE_MAX_INTEGER:
        raise _invalid_payload("Invalid purchase order payload")
    document_ids = payload["document_version_ids"]
    if not isinstance(document_ids, list):
        raise _invalid_payload("Invalid purchase order payload")
    normalized_documents = [
        _positive_integer(item, "Invalid purchase order payload")
        for item in document_ids
    ]
    if len(set(normalized_documents)) != len(normalized_documents):
        raise _invalid_payload("Invalid purchase order payload")
    expected_delivery = payload["expected_delivery_on"]
    if expected_delivery is not None:
        expected_delivery = _business_date(
            expected_delivery, "Invalid purchase order payload"
        )
    ordered_on = _business_date(payload["ordered_on"], "Invalid purchase order payload")
    if expected_delivery is not None and expected_delivery < ordered_on:
        raise _invalid_payload("Invalid purchase order payload")
    return {
        "order_no": _required_text(
            payload["order_no"], "Invalid purchase order payload"
        ),
        "supplier_company_id": _positive_integer(
            payload["supplier_company_id"], "Invalid purchase order payload"
        ),
        "ordered_on": ordered_on,
        "expected_delivery_on": expected_delivery,
        "lines": lines,
        "notes": _optional_text(payload["notes"], "Invalid purchase order payload"),
        "document_version_ids": normalized_documents,
    }


def _normalize_receipt(payload: dict[str, Any]) -> dict[str, object]:
    raw_lines = payload["lines"]
    if not isinstance(raw_lines, list) or not raw_lines:
        raise _invalid_payload("Invalid goods receipt payload")
    lines: list[dict[str, int]] = []
    seen: set[int] = set()
    for raw_line in raw_lines:
        if not isinstance(raw_line, dict) or set(raw_line) != _RECEIPT_LINE_FIELDS:
            raise _invalid_payload("Invalid goods receipt payload")
        line_id = _positive_integer(
            raw_line["purchase_order_line_id"], "Invalid goods receipt payload"
        )
        if line_id in seen:
            raise _invalid_payload("Invalid goods receipt payload")
        seen.add(line_id)
        lines.append(
            {
                "purchase_order_line_id": line_id,
                "quantity_milli": parse_quantity(raw_line["quantity"]),
            }
        )
    return {
        "received_on": _business_date(
            payload["received_on"], "Invalid goods receipt payload"
        ),
        "warehouse_name": _required_text(
            payload["warehouse_name"], "Invalid goods receipt payload"
        ),
        "lines": lines,
        "notes": _optional_text(payload["notes"], "Invalid goods receipt payload"),
    }


def _list_detail(
    connection: sqlite3.Connection,
    row: sqlite3.Row,
    project_code: str,
) -> dict[str, object]:
    lines = _list_line_rows(connection, int(row["id"]))
    responses = [
        _line_response(connection, line, int(row["revision"])) for line in lines
    ]
    return {
        "id": row["id"],
        "project_code": project_code,
        "name": row["name"],
        "notes": row["notes"],
        "status": row["status"],
        "revision": row["revision"],
        "confirmed_at": row["confirmed_at"],
        "line_count": len(responses),
        "cost_total_cents": sum(int(line["cost_total_cents"]) for line in responses),
        "quoted_total_cents": sum(
            int(line["quoted_total_cents"]) for line in responses
        ),
        "lines": responses,
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def _list_summary(row: sqlite3.Row, project_code: str) -> dict[str, object]:
    return {
        "id": row["id"],
        "project_code": project_code,
        "name": row["name"],
        "notes": row["notes"],
        "status": row["status"],
        "revision": row["revision"],
        "line_count": row["line_count"],
        "confirmed_at": row["confirmed_at"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def _line_response(
    connection: sqlite3.Connection,
    row: sqlite3.Row,
    list_revision: int | None,
) -> dict[str, object]:
    quantities = _line_facts(connection, int(row["id"]))
    required = int(row["quantity_milli"])
    ordered = int(quantities["ordered_quantity_milli"])
    received = int(quantities["received_quantity_milli"])
    used = int(quantities["used_quantity_milli"])
    ordered_amount = int(quantities["ordered_amount_cents"])
    paid_amount = int(quantities["paid_amount_cents"])
    invoiced_amount = int(quantities["invoiced_amount_cents"])
    response = {
        "id": row["id"],
        "procurement_list_id": row["procurement_list_id"],
        "inventory_item_id": row["inventory_item_id"],
        "sequence_no": row["sequence_no"],
        "category": row["category"],
        "name": row["name"],
        "specification": row["specification"],
        "brand": row["brand"],
        "model": row["model"],
        "quantity": format_quantity(required),
        "unit": row["unit"],
        "unit_cost_cents": row["unit_cost_cents"],
        "quoted_unit_price_cents": row["quoted_unit_price_cents"],
        "cost_total_cents": line_value(int(row["unit_cost_cents"]), required),
        "quoted_total_cents": line_value(int(row["quoted_unit_price_cents"]), required),
        "ordered_quantity": format_quantity(ordered),
        "ordered_amount_cents": ordered_amount,
        "paid_amount_cents": paid_amount,
        "received_quantity": format_quantity(received),
        "invoiced_amount_cents": invoiced_amount,
        "issued_quantity": format_quantity(used),
        "order_status": _quantity_status(
            ordered,
            required,
            zero="not_ordered",
            partial="partial",
            complete="ordered",
            over="over_ordered",
        ),
        "payment_status": _amount_status(
            paid_amount,
            ordered_amount,
            zero="unpaid",
            complete="paid",
        ),
        "receipt_status": _quantity_status(
            received,
            ordered,
            zero="not_received",
            partial="partial",
            complete="received",
            over="received",
        ),
        "invoice_status": _amount_status(
            invoiced_amount,
            ordered_amount,
            zero="not_invoiced",
            complete="invoiced",
        ),
        "usage_status": _quantity_status(
            used,
            required,
            zero="unused",
            partial="partial",
            complete="used",
            over="used",
        ),
        "revision": row["revision"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }
    if list_revision is not None:
        response["list_revision"] = list_revision
    return response


def _line_facts(connection: sqlite3.Connection, line_id: int) -> dict[str, int]:
    row = connection.execute(
        """
        SELECT
            COALESCE((
                SELECT SUM(order_lines.quantity_milli)
                FROM purchase_order_lines AS order_lines
                JOIN purchase_orders AS orders
                    ON orders.id = order_lines.purchase_order_id
                WHERE order_lines.procurement_line_id = procurement_lines.id
                  AND orders.status IN (
                      'confirmed', 'partially_received', 'received'
                  )
            ), 0) AS ordered_quantity_milli,
            COALESCE((
                SELECT SUM(
                    (order_lines.unit_cost_cents * order_lines.quantity_milli + 500)
                    / 1000
                )
                FROM purchase_order_lines AS order_lines
                JOIN purchase_orders AS orders
                    ON orders.id = order_lines.purchase_order_id
                WHERE order_lines.procurement_line_id = procurement_lines.id
                  AND orders.status IN (
                      'confirmed', 'partially_received', 'received'
                  )
            ), 0) AS ordered_amount_cents,
            COALESCE((
                SELECT SUM(receipt_lines.quantity_milli)
                FROM goods_receipt_lines AS receipt_lines
                JOIN goods_receipts AS receipts
                    ON receipts.id = receipt_lines.goods_receipt_id
                JOIN purchase_order_lines AS order_lines
                    ON order_lines.id = receipt_lines.purchase_order_line_id
                WHERE order_lines.procurement_line_id = procurement_lines.id
                  AND receipts.status = 'active'
            ), 0) AS received_quantity_milli,
            COALESCE((
                SELECT SUM(issue_lines.quantity_milli)
                FROM inventory_issue_lines AS issue_lines
                JOIN inventory_issues AS issues
                    ON issues.id = issue_lines.inventory_issue_id
                WHERE issue_lines.procurement_line_id = procurement_lines.id
                  AND issues.status = 'active'
            ), 0) AS used_quantity_milli,
            COALESCE((
                SELECT SUM(allocations.amount_cents)
                FROM supplier_payment_allocations AS allocations
                JOIN supplier_payments AS payments
                    ON payments.id = allocations.supplier_payment_id
                JOIN purchase_order_lines AS order_lines
                    ON order_lines.id = allocations.purchase_order_line_id
                WHERE order_lines.procurement_line_id = procurement_lines.id
                  AND payments.status = 'active'
            ), 0) AS paid_amount_cents,
            COALESCE((
                SELECT SUM(allocations.amount_cents)
                FROM supplier_invoice_allocations AS allocations
                JOIN supplier_invoices AS invoices
                    ON invoices.id = allocations.supplier_invoice_id
                JOIN purchase_order_lines AS order_lines
                    ON order_lines.id = allocations.purchase_order_line_id
                WHERE order_lines.procurement_line_id = procurement_lines.id
                  AND invoices.status = 'active'
            ), 0) AS invoiced_amount_cents
        FROM procurement_lines
        WHERE procurement_lines.id = ?
        """,
        (line_id,),
    ).fetchone()
    if row is None:
        raise sqlite3.DatabaseError("procurement line facts are missing")
    return {key: int(value) for key, value in dict(row).items()}


def _order_response(
    connection: sqlite3.Connection,
    row: sqlite3.Row,
    project_code: str,
) -> dict[str, object]:
    lines = _order_lines(connection, int(row["id"]))
    supplier = connection.execute(
        "SELECT name FROM companies WHERE id = ?",
        (row["supplier_company_id"],),
    ).fetchone()
    documents = connection.execute(
        """
        SELECT document_version_id FROM purchase_order_documents
        WHERE purchase_order_id = ? ORDER BY document_version_id
        """,
        (row["id"],),
    ).fetchall()
    ordered_amount = sum(
        line_value(int(line["unit_cost_cents"]), int(line["quantity_milli"]))
        for line in lines
    )
    return {
        "id": row["id"],
        "project_code": project_code,
        "order_no": row["order_no"],
        "supplier_company_id": row["supplier_company_id"],
        "supplier_company_name": None if supplier is None else supplier["name"],
        "ordered_on": row["ordered_on"],
        "expected_delivery_on": row["expected_delivery_on"],
        "notes": row["notes"],
        "status": row["status"],
        "ordered_amount_cents": ordered_amount,
        "document_version_ids": [document[0] for document in documents],
        "revision": row["revision"],
        "lines": [
            {
                "id": line["id"],
                "purchase_order_id": line["purchase_order_id"],
                "procurement_line_id": line["procurement_line_id"],
                "quantity": format_quantity(int(line["quantity_milli"])),
                "received_quantity": format_quantity(
                    int(line["received_quantity_milli"])
                ),
                "unit_cost_cents": line["unit_cost_cents"],
                "line_amount_cents": line_value(
                    int(line["unit_cost_cents"]), int(line["quantity_milli"])
                ),
                "overage_reason": line["overage_reason"],
            }
            for line in lines
        ],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def _receipt_response(
    connection: sqlite3.Connection,
    row: sqlite3.Row,
) -> dict[str, object]:
    lines = connection.execute(
        """
        SELECT * FROM goods_receipt_lines
        WHERE goods_receipt_id = ? ORDER BY id
        """,
        (row["id"],),
    ).fetchall()
    return {
        "id": row["id"],
        "purchase_order_id": row["purchase_order_id"],
        "received_on": row["received_on"],
        "warehouse_name": row["warehouse_name"],
        "notes": row["notes"],
        "status": row["status"],
        "revision": row["revision"],
        "lines": [
            {
                "id": line["id"],
                "purchase_order_line_id": line["purchase_order_line_id"],
                "inventory_item_id": line["inventory_item_id"],
                "quantity": format_quantity(int(line["quantity_milli"])),
                "value_cents": line["value_cents"],
                "movement_id": line["movement_id"],
            }
            for line in lines
        ],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def _refresh_order_receipt_status(
    connection: sqlite3.Connection,
    order_id: int,
    timestamp: str,
) -> None:
    totals = connection.execute(
        """
        SELECT SUM(quantity_milli) AS ordered,
               SUM(received_quantity_milli) AS received
        FROM purchase_order_lines WHERE purchase_order_id = ?
        """,
        (order_id,),
    ).fetchone()
    next_status = (
        "received" if totals["received"] == totals["ordered"] else "partially_received"
    )
    connection.execute(
        """
        UPDATE purchase_orders
        SET status = ?, revision = revision + 1, updated_at = ? WHERE id = ?
        """,
        (next_status, timestamp, order_id),
    )


def _validate_document_versions(
    connection: sqlite3.Connection,
    project_code: str,
    document_ids: list[int],
) -> None:
    if not document_ids:
        return
    placeholders = ",".join("?" for _ in document_ids)
    found = connection.execute(
        f"""
        SELECT versions.id
        FROM document_versions AS versions
        JOIN documents ON documents.id = versions.document_id
        WHERE versions.id IN ({placeholders})
          AND documents.project_code = ? COLLATE NOCASE
          AND documents.archived_at IS NULL
        """,
        (*document_ids, project_code),
    ).fetchall()
    if {row[0] for row in found} != set(document_ids):
        raise _invalid_payload("Invalid purchase order payload")


def _project(connection: sqlite3.Connection, project_key: str) -> sqlite3.Row:
    row = connection.execute(
        "SELECT * FROM projects WHERE project_code_key = ?",
        (project_key,),
    ).fetchone()
    if row is None:
        raise _not_found("Project not found")
    return row


def _active_project(connection: sqlite3.Connection, project_key: str) -> sqlite3.Row:
    row = _project(connection, project_key)
    if row["status"] != "active":
        raise _business_conflict("Project is archived", "PROJECT_ARCHIVED")
    return row


def _list_row(
    connection: sqlite3.Connection,
    list_id: int,
    project_id: int,
) -> sqlite3.Row | None:
    return connection.execute(
        "SELECT * FROM procurement_lists WHERE id = ? AND project_id = ?",
        (list_id, project_id),
    ).fetchone()


def _editable_list(
    connection: sqlite3.Connection,
    list_id: int,
    project_id: int,
) -> sqlite3.Row:
    row = _list_row(connection, list_id, project_id)
    if row is None:
        raise _not_found("Procurement list not found")
    if row["status"] != "draft":
        raise _business_conflict(
            "Procurement list is not editable",
            "PROCUREMENT_LIST_NOT_EDITABLE",
        )
    return row


def _line_row(
    connection: sqlite3.Connection,
    line_id: int,
    list_id: int,
) -> sqlite3.Row | None:
    return connection.execute(
        """
        SELECT * FROM procurement_lines
        WHERE id = ? AND procurement_list_id = ?
        """,
        (line_id, list_id),
    ).fetchone()


def _list_line_rows(connection: sqlite3.Connection, list_id: int) -> list[sqlite3.Row]:
    return connection.execute(
        """
        SELECT * FROM procurement_lines
        WHERE procurement_list_id = ? ORDER BY sequence_no, id
        """,
        (list_id,),
    ).fetchall()


def _project_line_rows(
    connection: sqlite3.Connection,
    project_id: int,
) -> list[sqlite3.Row]:
    return connection.execute(
        """
        SELECT lines.*
        FROM procurement_lines AS lines
        JOIN procurement_lists AS lists ON lists.id = lines.procurement_list_id
        WHERE lists.project_id = ? AND lists.status = 'confirmed'
        ORDER BY lists.id, lines.sequence_no, lines.id
        """,
        (project_id,),
    ).fetchall()


def _confirmed_procurement_line(
    connection: sqlite3.Connection,
    line_id: int,
    project_id: int,
) -> sqlite3.Row:
    row = connection.execute(
        """
        SELECT lines.*
        FROM procurement_lines AS lines
        JOIN procurement_lists AS lists ON lists.id = lines.procurement_list_id
        WHERE lines.id = ? AND lists.project_id = ? AND lists.status = 'confirmed'
        """,
        (line_id, project_id),
    ).fetchone()
    if row is None:
        raise _invalid_payload("Invalid purchase order payload")
    return row


def _order_row(
    connection: sqlite3.Connection,
    order_id: int,
    project_id: int,
) -> sqlite3.Row | None:
    return connection.execute(
        "SELECT * FROM purchase_orders WHERE id = ? AND project_id = ?",
        (order_id, project_id),
    ).fetchone()


def _order_lines(connection: sqlite3.Connection, order_id: int) -> list[sqlite3.Row]:
    return connection.execute(
        """
        SELECT * FROM purchase_order_lines
        WHERE purchase_order_id = ? ORDER BY id
        """,
        (order_id,),
    ).fetchall()


def _quantity_status(
    actual: int,
    target: int,
    *,
    zero: str,
    partial: str,
    complete: str,
    over: str,
) -> str:
    if actual == 0 or target == 0:
        return zero
    if actual < target:
        return partial
    if actual == target:
        return complete
    return over


def _amount_status(
    actual: int,
    target: int,
    *,
    zero: str,
    complete: str,
) -> str:
    if actual == 0 or target == 0:
        return zero
    return "partial" if actual < target else complete


async def _read_json(
    request: Request,
    expected_fields: set[str],
    detail: str,
) -> dict[str, Any]:
    try:
        payload: Any = await request.json()
    except (RecursionError, UnicodeError, ValueError):
        raise _invalid_payload(detail) from None
    if not isinstance(payload, dict) or set(payload) != expected_fields:
        raise _invalid_payload(detail)
    return payload


def _read_pagination(request: Request) -> tuple[int, int]:
    return (
        _query_integer(request, "page", default=1, maximum=_SQLITE_MAX_INTEGER),
        _query_integer(request, "page_size", default=50, maximum=_MAX_PAGE_SIZE),
    )


def _query_integer(request: Request, name: str, *, default: int, maximum: int) -> int:
    values = request.query_params.getlist(name)
    if not values:
        return default
    if len(values) != 1 or not values[0].isascii() or not values[0].isdecimal():
        raise _invalid_payload("Invalid pagination")
    value = int(values[0])
    if not 1 <= value <= maximum:
        raise _invalid_payload("Invalid pagination")
    return value


def _single_query(request: Request, name: str) -> str | None:
    values = request.query_params.getlist(name)
    if not values:
        return None
    if len(values) != 1:
        raise _invalid_payload(f"Invalid {name}")
    return values[0]


def _normalize_project_path(project_code: str) -> str:
    try:
        return project_code_identity(normalize_project_code(project_code))
    except (TypeError, UnicodeError, ValueError):
        raise _invalid_payload("Invalid project code") from None


def _required_text(value: object, detail: str) -> str:
    normalized = _optional_text(value, detail)
    if normalized is None:
        raise _invalid_payload(detail)
    return normalized


def _optional_text(value: object, detail: str) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str):
        raise _invalid_payload(detail)
    normalized = value.strip()
    if not normalized:
        return None
    if "\x00" in normalized:
        raise _invalid_payload(detail)
    return normalized


def _positive_integer(value: object, detail: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int):
        raise _invalid_payload(detail)
    if not 1 <= value <= _SQLITE_MAX_INTEGER:
        raise _invalid_payload(detail)
    return value


def _money(value: object, detail: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int):
        raise _invalid_payload(detail)
    if not 0 <= value <= 9_000_000_000_000:
        raise _invalid_payload(detail)
    return value


def _business_date(value: object, detail: str) -> str:
    if not isinstance(value, str):
        raise _invalid_payload(detail)
    try:
        parsed = date.fromisoformat(value)
    except ValueError:
        raise _invalid_payload(detail) from None
    if parsed.isoformat() != value:
        raise _invalid_payload(detail)
    return value


def _parse_identifier(value: str) -> int:
    if not value.isascii() or not value.isdecimal():
        raise _invalid_payload("Invalid identifier")
    return _positive_integer(int(value), "Invalid identifier")


def _validate_idempotency_key(value: str) -> str:
    try:
        parsed = UUID(value)
    except (AttributeError, ValueError):
        raise _invalid_payload("Invalid Idempotency-Key") from None
    canonical = str(parsed)
    if value.lower() != canonical:
        raise _invalid_payload("Invalid Idempotency-Key")
    return canonical


def _request_hash(payload: object) -> str:
    return hashlib.sha256(
        json.dumps(
            payload,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode()
    ).hexdigest()


def _require_same_hash(stored_hash: str | None, request_hash: str) -> None:
    if stored_hash != request_hash:
        raise _business_conflict("Idempotency key reused", "IDEMPOTENCY_KEY_REUSED")


def _require_revision(row: sqlite3.Row, expected_revision: int) -> None:
    if row["revision"] != expected_revision:
        raise ApiError(
            status.HTTP_409_CONFLICT,
            "Resource was modified",
            "REVISION_CONFLICT",
            current_revision=int(row["revision"]),
            headers={
                "X-Error-Code": "REVISION_CONFLICT",
                "X-Current-Revision": str(row["revision"]),
            },
        )


def _last_insert_id(cursor: sqlite3.Cursor) -> int:
    if cursor.lastrowid is None:
        raise sqlite3.DatabaseError("insert did not return an identifier")
    return cursor.lastrowid


def _timestamp(clock: Clock) -> str:
    value = clock()
    if value.tzinfo is None or value.utcoffset() is None:
        raise ValueError("clock must return an aware datetime")
    return value.astimezone(timezone.utc).isoformat()


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _is_unique_constraint(failure: sqlite3.IntegrityError) -> bool:
    return (
        getattr(failure, "sqlite_errorcode", None) == sqlite3.SQLITE_CONSTRAINT_UNIQUE
    )


def _invalid_payload(detail: str) -> ApiError:
    return ApiError(
        status.HTTP_422_UNPROCESSABLE_CONTENT,
        detail,
        "VALIDATION_ERROR",
    )


def _not_found(detail: str) -> ApiError:
    return ApiError(status.HTTP_404_NOT_FOUND, detail, "RESOURCE_NOT_FOUND")


def _business_conflict(detail: str, error_code: str) -> ApiError:
    return ApiError(
        status.HTTP_409_CONFLICT,
        detail,
        error_code,
        headers={"X-Error-Code": error_code},
    )
