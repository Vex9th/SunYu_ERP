from __future__ import annotations

import asyncio
import json
import sqlite3
from collections.abc import Iterator
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from fastapi import FastAPI, Request
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from starlette.datastructures import UploadFile

from backend.app.core.config import Settings
from backend.app.core.database import connect_database
from backend.app.core.migrations import apply_migrations
from backend.app.core.security import SESSION_COOKIE_NAME, create_session_token
from backend.app.core.storage_paths import project_code_identity
from backend.app.features import files, procurement, procurement_extensions
from backend.app.features.api_common import ApiError

PROJECT_ROOT = Path(__file__).resolve().parents[3]
NOW = datetime(2026, 8, 31, 2, 0, tzinfo=timezone.utc)


@dataclass(frozen=True)
class Harness:
    app: FastAPI
    database_path: Path
    settings: Settings
    project_code: str
    other_project_code: str
    customer_company_id: int
    supplier_company_id: int
    document_version_id: int

    @contextmanager
    def client(
        self,
        *,
        raise_server_exceptions: bool = True,
    ) -> Iterator[TestClient]:
        with TestClient(
            self.app,
            raise_server_exceptions=raise_server_exceptions,
        ) as client:
            client.cookies.set(
                SESSION_COOKIE_NAME,
                create_session_token(self.settings.session_secret),
            )
            yield client

    @contextmanager
    def connection(self) -> Iterator[sqlite3.Connection]:
        connection = connect_database(self.database_path)
        try:
            yield connection
        finally:
            connection.close()


def _insert_company(connection: sqlite3.Connection, name: str) -> int:
    cursor = connection.execute(
        """
        INSERT INTO companies
            (name, taxpayer_id, registered_address, registered_phone,
             bank_name, bank_account, notes, created_at, updated_at)
        VALUES (?, NULL, NULL, NULL, NULL, NULL, NULL, ?, ?)
        """,
        (name, NOW.isoformat(), NOW.isoformat()),
    )
    assert cursor.lastrowid is not None
    return cursor.lastrowid


def _insert_project(
    connection: sqlite3.Connection,
    company_id: int,
    project_code: str,
) -> None:
    connection.execute(
        """
        INSERT INTO projects
            (project_code, project_code_key, company_id, name, description,
             created_at, updated_at)
        VALUES (?, ?, ?, ?, NULL, ?, ?)
        """,
        (
            project_code,
            project_code_identity(project_code),
            company_id,
            f"{project_code} 测试项目",
            NOW.isoformat(),
            NOW.isoformat(),
        ),
    )


def _insert_document_version(
    connection: sqlite3.Connection,
    project_code: str,
) -> int:
    document = connection.execute(
        """
        INSERT INTO documents
            (project_code, category, logical_name, notes, archive_reason,
             archived_at, revision, created_at, updated_at)
        VALUES (?, 'procurement', '供应商发票', NULL, NULL, NULL, 1, ?, ?)
        """,
        (project_code, NOW.isoformat(), NOW.isoformat()),
    )
    assert document.lastrowid is not None
    version = connection.execute(
        """
        INSERT INTO document_versions
            (document_id, version_number, original_filename, content_type,
             stored_relative_path, size_bytes, sha256, notes, created_at)
        VALUES (?, 1, 'invoice.pdf', 'application/pdf', ?, 1, ?, NULL, ?)
        """,
        (
            int(document.lastrowid),
            f"{project_code}/{document.lastrowid}/invoice.pdf",
            "0" * 64,
            NOW.isoformat(),
        ),
    )
    assert version.lastrowid is not None
    return int(version.lastrowid)


def _build_harness(
    tmp_path: Path,
    *,
    migrations_dir: Path | None = None,
) -> Harness:
    database_path = tmp_path / "erp.sqlite3"
    connection = connect_database(database_path)
    try:
        apply_migrations(
            connection,
            migrations_dir or PROJECT_ROOT / "backend" / "migrations",
        )
        customer_id = _insert_company(connection, "客户公司")
        supplier_id = _insert_company(connection, "供应商公司")
        _insert_project(connection, customer_id, "P-2026-001")
        _insert_project(connection, customer_id, "P-2026-002")
        document_version_id = _insert_document_version(connection, "P-2026-001")
    finally:
        connection.close()

    settings = Settings(
        config_path=tmp_path / "config.json",
        data_dir=tmp_path,
        backup_dir=None,
        backup_interval_hours=24,
        backup_retention_days=30,
        host="127.0.0.1",
        port=8765,
        session_secret="procurement-extension-session-secret",
    )

    def get_connection() -> Iterator[sqlite3.Connection]:
        owned = connect_database(database_path)
        try:
            yield owned
        finally:
            owned.close()

    def get_settings() -> Settings:
        return settings

    app = FastAPI()
    app.include_router(
        procurement.create_procurement_router(
            get_connection,
            get_settings,
            clock=lambda: NOW,
        )
    )
    app.include_router(
        procurement_extensions.create_procurement_extensions_router(
            get_connection,
            get_settings,
            clock=lambda: NOW,
        )
    )
    return Harness(
        app,
        database_path,
        settings,
        "P-2026-001",
        "P-2026-002",
        customer_id,
        supplier_id,
        document_version_id,
    )


def _line_payload(*, sequence_no: int = 1, name: str = "接触器") -> dict[str, object]:
    return {
        "sequence_no": sequence_no,
        "category": "电气",
        "name": name,
        "specification": "AC220V",
        "brand": "施耐德",
        "model": "LC1D09",
        "quantity": "5.000",
        "unit": "PCS",
        "unit_cost_cents": 1000,
        "quoted_unit_price_cents": 1500,
    }


def _create_confirmed_list(
    client: TestClient,
    project_code: str,
    *,
    key_prefix: str = "10000000",
) -> dict[str, object]:
    created = client.post(
        f"/api/projects/{project_code}/procurement-lists",
        headers={"Idempotency-Key": f"{key_prefix}-0000-4000-8000-000000000001"},
        json={"name": "采购清单", "notes": None},
    )
    assert created.status_code == 201, created.text
    procurement_list = created.json()
    line = client.post(
        f"/api/projects/{project_code}/procurement-lists/{procurement_list['id']}/lines",
        headers={"Idempotency-Key": f"{key_prefix}-0000-4000-8000-000000000002"},
        json=_line_payload(),
    )
    assert line.status_code == 201, line.text
    confirmed = client.post(
        f"/api/projects/{project_code}/procurement-lists/{procurement_list['id']}/confirm",
        headers={"Idempotency-Key": f"{key_prefix}-0000-4000-8000-000000000003"},
        json={"expected_revision": line.json()["list_revision"]},
    )
    assert confirmed.status_code == 200, confirmed.text
    return confirmed.json()


def _create_confirmed_order(
    client: TestClient,
    harness: Harness,
) -> dict[str, object]:
    procurement_list = _create_confirmed_list(client, harness.project_code)
    order = client.post(
        f"/api/projects/{harness.project_code}/purchase-orders",
        headers={"Idempotency-Key": "20000000-0000-4000-8000-000000000001"},
        json={
            "order_no": "PO-001",
            "supplier_company_id": harness.supplier_company_id,
            "ordered_on": "2026-08-31",
            "expected_delivery_on": None,
            "lines": [
                {
                    "procurement_line_id": procurement_list["lines"][0]["id"],
                    "quantity": "5.000",
                    "unit_cost_cents": 1000,
                    "overage_reason": None,
                }
            ],
            "notes": None,
            "document_version_ids": [],
        },
    )
    assert order.status_code == 201, order.text
    confirmed = client.post(
        f"/api/projects/{harness.project_code}/purchase-orders/{order.json()['id']}/confirm",
        headers={"Idempotency-Key": "20000000-0000-4000-8000-000000000002"},
        json={"expected_revision": order.json()["revision"]},
    )
    assert confirmed.status_code == 200, confirmed.text
    return confirmed.json()


def _workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        [
            "序号",
            "大类",
            "名称",
            "规格",
            "品牌",
            "型号",
            "数量",
            "单位",
            "成本单价（元）",
            "报价单价（元）",
        ]
    )
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _replace_zip_member(content: bytes, name: str, replacement: bytes) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(content)) as source, ZipFile(
        output, "w", compression=ZIP_DEFLATED
    ) as target:
        for member in source.infolist():
            target.writestr(
                member,
                replacement if member.filename == name else source.read(member),
            )
    return output.getvalue()


def _compressed_archive(member_count: int, member_size: int) -> bytes:
    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
        for index in range(member_count):
            archive.writestr(f"xl/payload-{index}.xml", b"A" * member_size)
    return output.getvalue()


def test_import_preview_is_read_only_and_confirm_is_atomic_canonical_idempotent(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO inventory_items
                (brand, name, model, specification, unit, notes,
                 quantity_milli, inventory_value_cents, revision,
                 create_idempotency_key, create_request_hash,
                 created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, NULL, 0, 0, 1, ?, ?, ?, ?)
            """,
            (
                " 施耐德 ",
                "接触器",
                "lc1d09",
                "AC220V",
                "pcs",
                "import-match-key",
                "import-match-hash",
                NOW.isoformat(),
                NOW.isoformat(),
            ),
        )
        assert cursor.lastrowid is not None
        matching_inventory_id = cursor.lastrowid
    content = _workbook_bytes(
        [
            [1, "电气", "接触器", "AC220V", "施耐德", "LC1D09", "5.000", "PCS", "10.00", "15.00"],
            [2, "辅材", "线槽", None, None, None, "2.500", "M", "3.25", "5.50"],
        ]
    )
    with harness.client() as client:
        preview_headers = {
            "Idempotency-Key": "30000000-0000-4000-8000-000000000010"
        }
        preview = client.post(
            f"/api/projects/{harness.project_code.lower()}/procurement-imports/preview",
            headers=preview_headers,
            files={
                "file": (
                    "采购清单.xlsx",
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert preview.status_code == 201, preview.text
        body = preview.json()
        assert body["project_code"] == harness.project_code
        assert body["status"] == "preview"
        assert body["revision"] == 1
        assert body["errors"] == []
        assert body["rows"] == [
            _line_payload(),
            {
                "sequence_no": 2,
                "category": "辅材",
                "name": "线槽",
                "specification": None,
                "brand": None,
                "model": None,
                "quantity": "2.500",
                "unit": "M",
                "unit_cost_cents": 325,
                "quoted_unit_price_cents": 550,
            },
        ]
        with harness.connection() as connection:
            assert connection.execute("SELECT COUNT(*) FROM procurement_lists").fetchone()[0] == 0
            assert connection.execute("SELECT COUNT(*) FROM procurement_lines").fetchone()[0] == 0

        headers = {"Idempotency-Key": "30000000-0000-4000-8000-000000000001"}
        payload = {"list_name": " Excel 导入 ", "expected_revision": 1}
        confirmed = client.post(
            f"/api/projects/{harness.project_code.lower()}/procurement-imports/{body['id']}/confirm",
            headers=headers,
            json=payload,
        )
        duplicate = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers={
                "Idempotency-Key": "30000000-0000-4000-8000-000000000011"
            },
            files={"file": ("同一份清单.xlsx", content)},
        )
        assert duplicate.status_code == 409
        assert duplicate.json()["error_code"] == "PROCUREMENT_IMPORT_ALREADY_CONFIRMED"
        with harness.connection() as connection:
            connection.execute(
                """
                UPDATE projects
                SET status = 'archived', archived_at = ?, archive_reason = ?
                WHERE project_code_key = ?
                """,
                (
                    NOW.isoformat(),
                    "测试归档",
                    project_code_identity(harness.project_code),
                ),
            )
        preview_replay = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers=preview_headers,
            files={
                "file": (
                    "采购清单.xlsx",
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        preview_key_reuse = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers=preview_headers,
            files={"file": ("另一份.xlsx", _workbook_bytes([[1, "电气", "不同物料", None, None, None, "1", "PCS", "1", "2"]]))},
        )
        replay = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/{body['id']}/confirm",
            headers=headers,
            json=payload,
        )

    assert confirmed.status_code == replay.status_code == 201
    assert confirmed.json() == replay.json()
    assert preview_replay.status_code == 201
    assert preview_replay.json() == preview.json()
    assert preview_key_reuse.status_code == 409
    assert preview_key_reuse.json()["error_code"] == "IDEMPOTENCY_KEY_REUSED"
    imported_list = confirmed.json()["procurement_list"]
    assert imported_list["name"] == "Excel 导入"
    assert imported_list["status"] == "draft"
    assert imported_list["line_count"] == 2
    assert imported_list["lines"][0]["inventory_item_id"] == matching_inventory_id
    assert imported_list["lines"][1]["inventory_item_id"] is None
    with harness.connection() as connection:
        assert connection.execute("SELECT COUNT(*) FROM procurement_lists").fetchone()[0] == 1
        assert connection.execute("SELECT COUNT(*) FROM procurement_lines").fetchone()[0] == 2


def test_import_reports_cell_errors_enforces_limits_and_never_partially_confirms(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    invalid = _workbook_bytes(
        [[1, "电气", None, "AC220V", None, None, "abc", "PCS", "-1.00", "15.001"]]
    )
    with harness.client() as client:
        anonymous = TestClient(harness.app).post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            files={"file": ("bad.xlsx", b"not-xlsx")},
        )
        assert anonymous.status_code == 401

        too_large = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers={
                "Idempotency-Key": "31000000-0000-4000-8000-000000000001"
            },
            files={"file": ("huge.xlsx", b"x" * (20 * 1024 * 1024 + 1))},
        )
        assert too_large.status_code == 422

        preview = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers={
                "Idempotency-Key": "31000000-0000-4000-8000-000000000002"
            },
            files={"file": ("invalid.xlsx", invalid)},
        )
        assert preview.status_code == 201, preview.text
        errors = preview.json()["errors"]
        assert {(item["row"], item["column"], item["field"]) for item in errors} == {
            (2, 3, "name"),
            (2, 7, "quantity"),
            (2, 9, "unit_cost_cents"),
            (2, 10, "quoted_unit_price_cents"),
        }
        confirm = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/{preview.json()['id']}/confirm",
            headers={"Idempotency-Key": "30000000-0000-4000-8000-000000000002"},
            json={"list_name": "不得创建", "expected_revision": 1},
        )
    assert confirm.status_code == 409
    assert confirm.json()["error_code"] == "PROCUREMENT_IMPORT_HAS_ERRORS"
    with harness.connection() as connection:
        assert connection.execute("SELECT COUNT(*) FROM procurement_lists").fetchone()[0] == 0
        assert connection.execute("SELECT COUNT(*) FROM procurement_lines").fetchone()[0] == 0

    too_many_rows = [[index, "电气", "物料", None, None, None, "1", "PCS", "1", "2"] for index in range(1, 10_002)]
    with harness.client() as client:
        response = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers={
                "Idempotency-Key": "31000000-0000-4000-8000-000000000003"
            },
            files={"file": ("too-many.xlsx", _workbook_bytes(too_many_rows))},
        )
    assert response.status_code == 422


def test_import_reports_duplicate_normalized_business_rows_and_blocks_confirm(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    content = _workbook_bytes(
        [
            [
                1,
                "电气",
                "接触器",
                "AC220V",
                "施耐德",
                "LC1D09",
                "5.000",
                "PCS",
                "10.00",
                "15.00",
            ],
            [
                2,
                " 电气 ",
                "接触器",
                "AC220V",
                "施耐德",
                "LC1D09",
                "5",
                " PCS ",
                "10",
                "15",
            ],
        ]
    )
    with harness.client() as client:
        preview = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers={
                "Idempotency-Key": "31000000-0000-4000-8000-000000000004"
            },
            files={"file": ("duplicate-lines.xlsx", content)},
        )
        assert preview.status_code == 201, preview.text
        assert preview.json()["rows"] == [_line_payload()]
        assert preview.json()["errors"] == [
            {
                "row": 3,
                "column": 0,
                "field": "row",
                "message": "业务字段重复",
            }
        ]

        confirm = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/"
            f"{preview.json()['id']}/confirm",
            headers={
                "Idempotency-Key": "31000000-0000-4000-8000-000000000005"
            },
            json={"list_name": "不得创建", "expected_revision": 1},
        )

    assert confirm.status_code == 409
    assert confirm.json()["error_code"] == "PROCUREMENT_IMPORT_HAS_ERRORS"
    with harness.connection() as connection:
        assert connection.execute("SELECT COUNT(*) FROM procurement_lists").fetchone()[0] == 0
        assert connection.execute("SELECT COUNT(*) FROM procurement_lines").fetchone()[0] == 0


def test_import_rejects_formula_malformed_ooxml_and_zip_bombs(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    harness = _build_harness(tmp_path)
    formula = _workbook_bytes(
        [[1, "电气", "公式物料", None, None, None, "=1+1", "PCS", "1", "2"]]
    )
    malformed = _replace_zip_member(
        _workbook_bytes(
            [[1, "电气", "畸形物料", None, None, None, "1", "PCS", "1", "2"]]
        ),
        "xl/worksheets/sheet1.xml",
        b"<worksheet><broken>",
    )
    with harness.client() as client:
        formula_response = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers={
                "Idempotency-Key": "32000000-0000-4000-8000-000000000001"
            },
            files={"file": ("formula.xlsx", formula)},
        )
        assert formula_response.status_code == 201
        assert formula_response.json()["rows"] == []
        assert [error["field"] for error in formula_response.json()["errors"]] == [
            "quantity"
        ]

        malformed_response = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers={
                "Idempotency-Key": "32000000-0000-4000-8000-000000000002"
            },
            files={"file": ("malformed.xlsx", malformed)},
        )
        assert malformed_response.status_code == 422
        assert malformed_response.json()["error_code"] == "VALIDATION_ERROR"

        def must_not_parse(_: BytesIO, **__: object) -> Workbook:
            raise AssertionError("unsafe archive reached openpyxl")

        monkeypatch.setattr(procurement_extensions, "load_workbook", must_not_parse)
        ratio_bomb = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers={
                "Idempotency-Key": "32000000-0000-4000-8000-000000000003"
            },
            files={"file": ("ratio-bomb.xlsx", _compressed_archive(1, 2_000_000))},
        )
        assert ratio_bomb.status_code == 422
        member_bomb = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers={
                "Idempotency-Key": "32000000-0000-4000-8000-000000000004"
            },
            files={"file": ("member-bomb.xlsx", _compressed_archive(300, 1))},
        )
        assert member_bomb.status_code == 422


def test_import_limits_asgi_receive_and_closes_every_multipart_upload(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    boundary = b"procurement-boundary"
    prefix = (
        b"--"
        + boundary
        + b'\r\nContent-Disposition: form-data; name="file"; filename="large.xlsx"'
        + b"\r\nContent-Type: application/octet-stream\r\n\r\n"
    )
    chunks = [prefix, *([b"x" * (1024 * 1024)] * 22), b"\r\n--" + boundary + b"--\r\n"]
    received = 0

    async def receive() -> dict[str, object]:
        nonlocal received
        if received >= len(chunks):
            return {"type": "http.disconnect"}
        body = chunks[received]
        received += 1
        return {
            "type": "http.request",
            "body": body,
            "more_body": received < len(chunks),
        }

    request = Request(
        {
            "type": "http",
            "asgi": {"version": "3.0"},
            "http_version": "1.1",
            "method": "POST",
            "scheme": "http",
            "path": "/",
            "raw_path": b"/",
            "query_string": b"",
            "headers": [
                (
                    b"content-type",
                    b"multipart/form-data; boundary=" + boundary,
                )
            ],
            "client": ("test", 123),
            "server": ("test", 80),
        },
        receive,
    )
    with pytest.raises(ApiError) as failure:
        asyncio.run(procurement_extensions._read_xlsx_upload(request))
    assert failure.value.status_code == 422
    assert received < len(chunks)

    harness = _build_harness(tmp_path)
    closed: list[str | None] = []
    original_close = UploadFile.close

    async def tracked_close(upload: UploadFile) -> None:
        closed.append(upload.filename)
        await original_close(upload)

    monkeypatch.setattr(UploadFile, "close", tracked_close)
    with harness.client() as client:
        response = client.post(
            f"/api/projects/{harness.project_code}/procurement-imports/preview",
            headers={
                "Idempotency-Key": "33000000-0000-4000-8000-000000000001"
            },
            files=[
                ("file", ("first.xlsx", _workbook_bytes([]))),
                ("extra", ("second.xlsx", _workbook_bytes([]))),
            ],
        )
    assert response.status_code == 422
    assert sorted(closed) == ["first.xlsx", "second.xlsx"]


def test_purchase_order_multipart_creates_managed_contract_documents_and_replays(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        procurement_list = _create_confirmed_list(client, harness.project_code)
        payload = {
            "order_no": " PO-ATTACH-001 ",
            "supplier_company_id": harness.supplier_company_id,
            "ordered_on": "2026-08-31",
            "expected_delivery_on": None,
            "lines": [
                {
                    "procurement_line_id": procurement_list["lines"][0]["id"],
                    "quantity": "2.000",
                    "unit_cost_cents": 900,
                    "overage_reason": None,
                }
            ],
            "notes": "供应商合同随采购单归档",
            "document_version_ids": [harness.document_version_id],
        }
        headers = {
            "Idempotency-Key": "34000000-0000-4000-8000-000000000001"
        }
        uploads = [
            ("files", ("原始合同.pdf", b"contract", "application/pdf")),
            ("files", ("盖章页.jpg", b"signed", "image/jpeg")),
        ]
        endpoint = f"/api/projects/{harness.project_code}/purchase-orders"
        created = client.post(
            endpoint,
            headers=headers,
            data={"payload": json.dumps(payload, ensure_ascii=False)},
            files=uploads,
        )
        replay = client.post(
            f"/api/projects/{harness.project_code.lower()}/purchase-orders",
            headers=headers,
            data={"payload": json.dumps(payload, ensure_ascii=False)},
            files=uploads,
        )
        reused = client.post(
            endpoint,
            headers=headers,
            data={"payload": json.dumps(payload, ensure_ascii=False)},
            files=[
                ("files", ("原始合同.pdf", b"changed", "application/pdf")),
                ("files", ("盖章页.jpg", b"signed", "image/jpeg")),
            ],
        )

    assert created.status_code == replay.status_code == 201, created.text
    assert replay.json() == created.json()
    assert reused.status_code == 409
    assert reused.json()["error_code"] == "IDEMPOTENCY_KEY_REUSED"
    assert created.json()["document_version_ids"][0] == harness.document_version_id
    uploaded_version_ids = created.json()["document_version_ids"][1:]
    assert len(uploaded_version_ids) == 2
    with harness.connection() as connection:
        rows = connection.execute(
            """
            SELECT versions.original_filename, versions.managed_filename,
                   versions.stored_relative_path, documents.category,
                   documents.logical_name
            FROM document_versions AS versions
            JOIN documents ON documents.id = versions.document_id
            WHERE versions.id IN (?, ?)
            ORDER BY versions.id
            """,
            uploaded_version_ids,
        ).fetchall()
        assert connection.execute(
            "SELECT COUNT(*) FROM purchase_orders"
        ).fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM purchase_order_documents WHERE purchase_order_id = ?",
            (created.json()["id"],),
        ).fetchone()[0] == 3
    assert [row["managed_filename"] for row in rows] == [
        "P-2026-001_采购合同_PO-ATTACH-001_20260831_01.pdf",
        "P-2026-001_采购合同_PO-ATTACH-001_20260831_02.jpg",
    ]
    assert [row["original_filename"] for row in rows] == ["原始合同.pdf", "盖章页.jpg"]
    assert all(row["category"] == "procurement_contract" for row in rows)
    assert all("PO-ATTACH-001" in row["logical_name"] for row in rows)
    for row in rows:
        stored = harness.settings.data_dir / row["stored_relative_path"]
        assert stored.is_file()
        assert stored.name == row["managed_filename"]


def test_purchase_order_json_and_zero_file_multipart_share_idempotency(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        procurement_list = _create_confirmed_list(client, harness.project_code)
        payload = {
            "order_no": "PO-JSON-COMPAT",
            "supplier_company_id": harness.supplier_company_id,
            "ordered_on": "2026-08-31",
            "expected_delivery_on": None,
            "lines": [
                {
                    "procurement_line_id": procurement_list["lines"][0]["id"],
                    "quantity": "1.000",
                    "unit_cost_cents": 900,
                    "overage_reason": None,
                }
            ],
            "notes": None,
            "document_version_ids": [],
        }
        endpoint = f"/api/projects/{harness.project_code}/purchase-orders"
        headers = {
            "Idempotency-Key": "34100000-0000-4000-8000-000000000001"
        }
        created = client.post(endpoint, headers=headers, json=payload)
        replay = client.post(
            endpoint,
            headers=headers,
            files={
                "payload": (
                    None,
                    json.dumps(payload, ensure_ascii=False),
                    "application/json",
                )
            },
        )

    assert created.status_code == replay.status_code == 201, replay.text
    assert replay.json() == created.json()
    with harness.connection() as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM purchase_orders"
        ).fetchone()[0] == 1


def test_purchase_order_attachment_database_failure_rolls_back_files_and_order(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        procurement_list = _create_confirmed_list(client, harness.project_code)
    with harness.connection() as connection:
        connection.execute(
            """
            CREATE TRIGGER reject_purchase_order_attachment_version
            BEFORE INSERT ON document_versions
            BEGIN
                SELECT RAISE(ABORT, 'injected purchase attachment database failure');
            END
            """
        )
    payload = {
        "order_no": "PO-DB-FAIL",
        "supplier_company_id": harness.supplier_company_id,
        "ordered_on": "2026-08-31",
        "expected_delivery_on": None,
        "lines": [
            {
                "procurement_line_id": procurement_list["lines"][0]["id"],
                "quantity": "1.000",
                "unit_cost_cents": 900,
                "overage_reason": None,
            }
        ],
        "notes": None,
        "document_version_ids": [],
    }
    with harness.client(raise_server_exceptions=False) as client:
        response = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders",
            headers={
                "Idempotency-Key": "34200000-0000-4000-8000-000000000001"
            },
            data={"payload": json.dumps(payload)},
            files={"files": ("contract.pdf", b"contract", "application/pdf")},
        )

    assert response.status_code == 500
    with harness.connection() as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM purchase_orders"
        ).fetchone()[0] == 0
        assert connection.execute("SELECT COUNT(*) FROM documents").fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM document_versions"
        ).fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM purchase_order_documents"
        ).fetchone()[0] == 0
    assert not [
        path
        for path in (harness.settings.data_dir / "Projects").rglob("*")
        if path.is_file()
    ]
    assert not list((harness.settings.data_dir / "Temp").glob(".upload-*.tmp"))


def test_purchase_order_second_attachment_publish_failure_rolls_back_everything(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        procurement_list = _create_confirmed_list(client, harness.project_code)
    original_publish = files.publish_staged_version
    calls = 0

    def fail_second_publish(*args: object, **kwargs: object):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise OSError("injected purchase attachment publish failure")
        return original_publish(*args, **kwargs)

    monkeypatch.setattr(files, "publish_staged_version", fail_second_publish)
    payload = {
        "order_no": "PO-FS-FAIL",
        "supplier_company_id": harness.supplier_company_id,
        "ordered_on": "2026-08-31",
        "expected_delivery_on": None,
        "lines": [
            {
                "procurement_line_id": procurement_list["lines"][0]["id"],
                "quantity": "1.000",
                "unit_cost_cents": 900,
                "overage_reason": None,
            }
        ],
        "notes": None,
        "document_version_ids": [],
    }
    with harness.client(raise_server_exceptions=False) as client:
        response = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders",
            headers={
                "Idempotency-Key": "34300000-0000-4000-8000-000000000001"
            },
            data={"payload": json.dumps(payload)},
            files=[
                ("files", ("first.pdf", b"first", "application/pdf")),
                ("files", ("second.pdf", b"second", "application/pdf")),
            ],
        )

    assert response.status_code == 500
    with harness.connection() as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM purchase_orders"
        ).fetchone()[0] == 0
        assert connection.execute("SELECT COUNT(*) FROM documents").fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM document_versions"
        ).fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM purchase_order_documents"
        ).fetchone()[0] == 0
    assert not [
        path
        for path in (harness.settings.data_dir / "Projects").rglob("*")
        if path.is_file()
    ]
    assert not list((harness.settings.data_dir / "Temp").glob(".upload-*.tmp"))


def test_draft_order_update_and_cancel_are_strict_revisioned_and_idempotent(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        procurement_list = _create_confirmed_list(client, harness.project_code)
        line_id = procurement_list["lines"][0]["id"]
        created = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders",
            headers={"Idempotency-Key": "40000000-0000-4000-8000-000000000001"},
            json={
                "order_no": " PO-EDIT ",
                "supplier_company_id": harness.supplier_company_id,
                "ordered_on": "2026-08-31",
                "expected_delivery_on": None,
                "lines": [{"procurement_line_id": line_id, "quantity": "2.000", "unit_cost_cents": 900, "overage_reason": None}],
                "notes": None,
                "document_version_ids": [],
            },
        )
        assert created.status_code == 201, created.text
        update_payload = {
            "order_no": "PO-UPDATED",
            "supplier_company_id": harness.supplier_company_id,
            "ordered_on": "2026-08-31",
            "expected_delivery_on": "2026-09-03",
            "lines": [{"procurement_line_id": line_id, "quantity": "3.000", "unit_cost_cents": 950, "overage_reason": None}],
            "notes": "改价",
            "document_version_ids": [],
            "expected_revision": created.json()["revision"],
        }
        updated = client.put(
            f"/api/projects/{harness.project_code}/purchase-orders/{created.json()['id']}",
            json=update_payload,
        )
        assert updated.status_code == 200, updated.text
        assert updated.json()["order_no"] == "PO-UPDATED"
        assert updated.json()["lines"][0]["quantity"] == "3.000"
        stale = client.put(
            f"/api/projects/{harness.project_code}/purchase-orders/{created.json()['id']}",
            json=update_payload,
        )
        assert stale.status_code == 409
        assert stale.json()["error_code"] == "REVISION_CONFLICT"
        extra = client.put(
            f"/api/projects/{harness.project_code}/purchase-orders/{created.json()['id']}",
            json={**update_payload, "extra": True},
        )
        assert extra.status_code == 422
        cross_project = client.put(
            f"/api/projects/{harness.other_project_code}/purchase-orders/{created.json()['id']}",
            json={**update_payload, "expected_revision": updated.json()["revision"]},
        )
        assert cross_project.status_code == 404

        headers = {"Idempotency-Key": "40000000-0000-4000-8000-000000000002"}
        cancel_payload = {"reason": "需求取消", "expected_revision": updated.json()["revision"]}
        cancelled = client.post(
            f"/api/projects/{harness.project_code.lower()}/purchase-orders/{created.json()['id']}/cancel",
            headers=headers,
            json=cancel_payload,
        )
        assert cancelled.status_code == 200, cancelled.text
        with harness.connection() as connection:
            connection.execute(
                """
                UPDATE projects
                SET status = 'archived', archived_at = ?, archive_reason = ?
                WHERE project_code_key = ?
                """,
                (
                    NOW.isoformat(),
                    "测试归档",
                    project_code_identity(harness.project_code),
                ),
            )
        replay = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{created.json()['id']}/cancel",
            headers=headers,
            json=cancel_payload,
        )
        reused = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{created.json()['id']}/cancel",
            headers=headers,
            json={**cancel_payload, "reason": "另一原因"},
        )
    assert replay.status_code == 200
    assert replay.json() == cancelled.json()
    assert replay.json()["status"] == "cancelled"
    assert reused.status_code == 409
    assert reused.json()["error_code"] == "IDEMPOTENCY_KEY_REUSED"
    with harness.client() as restarted_client:
        reloaded = restarted_client.get(
            f"/api/projects/{harness.project_code}/purchase-orders/{created.json()['id']}"
        )
    assert reloaded.status_code == 200
    assert reloaded.json()["cancel_reason"] == "需求取消"
    assert reloaded.json()["cancelled_at"] == NOW.isoformat()


def test_supplier_payments_and_invoices_are_exact_capped_and_reversible(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
        line_id = order["lines"][0]["id"]
        payment_payload = {
            "paid_on": "2026-08-31",
            "amount_cents": 2000,
            "payment_method": "银行转账",
            "reference_no": "PAY-001",
            "allocations": [{"purchase_order_line_id": line_id, "amount_cents": 2000}],
            "notes": None,
        }
        payment_headers = {"Idempotency-Key": "50000000-0000-4000-8000-000000000001"}
        payment = client.post(
            f"/api/projects/{harness.project_code.lower()}/purchase-orders/{order['id']}/supplier-payments",
            headers=payment_headers,
            json=payment_payload,
        )
        replay = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/supplier-payments",
            headers=payment_headers,
            json=payment_payload,
        )
        assert payment.status_code == replay.status_code == 201
        assert replay.json() == payment.json()
        assert payment.json()["allocations"] == payment_payload["allocations"]

        bad_sum = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/supplier-payments",
            headers={"Idempotency-Key": "50000000-0000-4000-8000-000000000002"},
            json={**payment_payload, "amount_cents": 1999},
        )
        assert bad_sum.status_code == 422
        over = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/supplier-payments",
            headers={"Idempotency-Key": "50000000-0000-4000-8000-000000000003"},
            json={**payment_payload, "amount_cents": 3001, "allocations": [{"purchase_order_line_id": line_id, "amount_cents": 3001}]},
        )
        assert over.status_code == 409
        assert over.json()["error_code"] == "PAYMENT_AMOUNT_EXCEEDED"
        wrong_order_line = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/supplier-payments",
            headers={"Idempotency-Key": "50000000-0000-4000-8000-000000000010"},
            json={
                **payment_payload,
                "amount_cents": 1,
                "allocations": [
                    {
                        "purchase_order_line_id": line_id + 999,
                        "amount_cents": 1,
                    }
                ],
            },
        )
        assert wrong_order_line.status_code == 422
        cancel_with_facts = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/cancel",
            headers={"Idempotency-Key": "50000000-0000-4000-8000-000000000011"},
            json={"reason": "已有付款", "expected_revision": order["revision"]},
        )
        assert cancel_with_facts.status_code == 409
        assert cancel_with_facts.json()["error_code"] == "PURCHASE_ORDER_HAS_ACTIVE_FACTS"

        reverse_headers = {"Idempotency-Key": "50000000-0000-4000-8000-000000000004"}
        reverse_payload = {"reason": "重复付款", "expected_revision": payment.json()["revision"]}
        reversed_payment = client.post(
            f"/api/projects/{harness.project_code.lower()}/supplier-payments/{payment.json()['id']}/reverse",
            headers=reverse_headers,
            json=reverse_payload,
        )
        reverse_replay = client.post(
            f"/api/projects/{harness.project_code}/supplier-payments/{payment.json()['id']}/reverse",
            headers=reverse_headers,
            json=reverse_payload,
        )
        assert reversed_payment.status_code == reverse_replay.status_code == 200
        assert reverse_replay.json() == reversed_payment.json()
        assert reversed_payment.json()["reversal_reason"] == "重复付款"
        assert reversed_payment.json()["reversed_at"] == NOW.isoformat()
        duplicate_reverse = client.post(
            f"/api/projects/{harness.project_code}/supplier-payments/{payment.json()['id']}/reverse",
            headers={"Idempotency-Key": "50000000-0000-4000-8000-000000000005"},
            json={"reason": "再次冲销", "expected_revision": reversed_payment.json()["revision"]},
        )
        assert duplicate_reverse.status_code == 409
        assert duplicate_reverse.json()["error_code"] == "SUPPLIER_PAYMENT_ALREADY_REVERSED"

        invoice_payload = {
            "invoice_no": "INV-001",
            "invoiced_on": "2026-08-31",
            "amount_cents": 5000,
            "allocations": [{"purchase_order_line_id": line_id, "amount_cents": 5000}],
            "document_version_ids": [harness.document_version_id],
        }
        invoice = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/supplier-invoices",
            headers={"Idempotency-Key": "50000000-0000-4000-8000-000000000006"},
            json=invoice_payload,
        )
        assert invoice.status_code == 201, invoice.text
        invoice_replay = client.post(
            f"/api/projects/{harness.project_code.lower()}/purchase-orders/{order['id']}/supplier-invoices",
            headers={"Idempotency-Key": "50000000-0000-4000-8000-000000000006"},
            json=invoice_payload,
        )
        assert invoice_replay.status_code == 201
        assert invoice_replay.json() == invoice.json()
        over_invoice = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/supplier-invoices",
            headers={"Idempotency-Key": "50000000-0000-4000-8000-000000000007"},
            json={**invoice_payload, "invoice_no": "INV-002", "amount_cents": 1, "allocations": [{"purchase_order_line_id": line_id, "amount_cents": 1}]},
        )
        assert over_invoice.status_code == 409
        assert over_invoice.json()["error_code"] == "INVOICE_AMOUNT_EXCEEDED"
        reversed_invoice = client.post(
            f"/api/projects/{harness.project_code}/supplier-invoices/{invoice.json()['id']}/reverse",
            headers={"Idempotency-Key": "50000000-0000-4000-8000-000000000008"},
            json={"reason": "发票作废", "expected_revision": invoice.json()["revision"]},
        )
        assert reversed_invoice.status_code == 200
        assert reversed_invoice.json()["status"] == "reversed"
        assert reversed_invoice.json()["document_version_ids"] == [
            harness.document_version_id
        ]
        assert reversed_invoice.json()["reversal_reason"] == "发票作废"
        assert reversed_invoice.json()["reversed_at"] == NOW.isoformat()
        reverse_replay = client.post(
            f"/api/projects/{harness.project_code.lower()}/supplier-invoices/{invoice.json()['id']}/reverse",
            headers={"Idempotency-Key": "50000000-0000-4000-8000-000000000008"},
            json={"reason": "发票作废", "expected_revision": invoice.json()["revision"]},
        )
        assert reverse_replay.status_code == 200
        assert reverse_replay.json() == reversed_invoice.json()
        duplicate_invoice_reverse = client.post(
            f"/api/projects/{harness.project_code}/supplier-invoices/{invoice.json()['id']}/reverse",
            headers={"Idempotency-Key": "50000000-0000-4000-8000-000000000009"},
            json={
                "reason": "再次冲销",
                "expected_revision": reversed_invoice.json()["revision"],
            },
        )
        assert duplicate_invoice_reverse.status_code == 409
        assert (
            duplicate_invoice_reverse.json()["error_code"]
            == "SUPPLIER_INVOICE_ALREADY_REVERSED"
        )


def test_supplier_invoice_active_number_uniqueness_migrates_existing_history(
    tmp_path: Path,
) -> None:
    legacy_migrations = tmp_path / "legacy-migrations"
    legacy_migrations.mkdir()
    for source in sorted((PROJECT_ROOT / "backend" / "migrations").glob("*.sql")):
        if source.name >= "022_":
            continue
        (legacy_migrations / source.name).write_text(
            source.read_text(encoding="utf-8"),
            encoding="utf-8",
        )

    harness = _build_harness(tmp_path, migrations_dir=legacy_migrations)
    invoice_payload: dict[str, object]
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
        line_id = order["lines"][0]["id"]
        invoice_payload = {
            "invoice_no": "INV-REUSABLE-001",
            "invoiced_on": "2026-08-31",
            "amount_cents": 1000,
            "allocations": [
                {"purchase_order_line_id": line_id, "amount_cents": 1000}
            ],
            "document_version_ids": [harness.document_version_id],
        }
        original = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/"
            f"{order['id']}/supplier-invoices",
            headers={
                "Idempotency-Key": "73600000-0000-4000-8000-000000000001"
            },
            json=invoice_payload,
        )
        assert original.status_code == 201, original.text
        reversed_invoice = client.post(
            f"/api/projects/{harness.project_code}/supplier-invoices/"
            f"{original.json()['id']}/reverse",
            headers={
                "Idempotency-Key": "73600000-0000-4000-8000-000000000002"
            },
            json={"reason": "原票作废", "expected_revision": 1},
        )
        assert reversed_invoice.status_code == 200, reversed_invoice.text

    with harness.connection() as connection:
        migration_022 = (
            PROJECT_ROOT
            / "backend"
            / "migrations"
            / "022_supplier_invoice_active_number.sql"
        )
        (legacy_migrations / migration_022.name).write_text(
            migration_022.read_text(encoding="utf-8"),
            encoding="utf-8",
        )
        assert apply_migrations(
            connection,
            legacy_migrations,
        ) == ["022_supplier_invoice_active_number"]
        assert connection.execute("PRAGMA foreign_key_check").fetchall() == []

    with harness.client() as client:
        replacement = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/"
            f"{order['id']}/supplier-invoices",
            headers={
                "Idempotency-Key": "73600000-0000-4000-8000-000000000003"
            },
            json={**invoice_payload, "document_version_ids": []},
        )
        active_duplicate = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/"
            f"{order['id']}/supplier-invoices",
            headers={
                "Idempotency-Key": "73600000-0000-4000-8000-000000000004"
            },
            json={**invoice_payload, "document_version_ids": []},
        )

    assert replacement.status_code == 201, replacement.text
    assert replacement.json()["id"] != original.json()["id"]
    assert active_duplicate.status_code == 409
    assert active_duplicate.json()["error_code"] == "SUPPLIER_INVOICE_EXISTS"
    with harness.connection() as connection:
        invoice_rows = connection.execute(
            """
            SELECT id, status, reversal_reason, reversed_at
            FROM supplier_invoices
            WHERE purchase_order_id = ? AND invoice_no = ?
            ORDER BY id
            """,
            (order["id"], invoice_payload["invoice_no"]),
        ).fetchall()
        assert [row["status"] for row in invoice_rows] == ["reversed", "active"]
        assert invoice_rows[0]["reversal_reason"] == "原票作废"
        assert invoice_rows[0]["reversed_at"] == NOW.isoformat()
        assert connection.execute(
            """
            SELECT COUNT(*)
            FROM supplier_invoice_allocations
            WHERE supplier_invoice_id = ?
            """,
            (original.json()["id"],),
        ).fetchone()[0] == 1
        assert connection.execute(
            """
            SELECT document_version_id
            FROM supplier_invoice_documents
            WHERE supplier_invoice_id = ?
            """,
            (original.json()["id"],),
        ).fetchone()[0] == harness.document_version_id


def test_supplier_invoice_multipart_creates_managed_documents_and_replays(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
        line_id = order["lines"][0]["id"]
        payload = {
            "invoice_no": "PUR-INV-001",
            "invoiced_on": "2026-08-31",
            "amount_cents": 5000,
            "allocations": [
                {"purchase_order_line_id": line_id, "amount_cents": 5000}
            ],
            "document_version_ids": [],
        }
        key = "73000000-0000-4000-8000-000000000001"
        uploads = [
            ("files", ("供应商原票.pdf", b"supplier-invoice", "application/pdf")),
            ("files", ("抵扣联.png", b"tax-copy", "image/png")),
        ]
        created = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/"
            f"{order['id']}/supplier-invoices",
            headers={"Idempotency-Key": key},
            data={"payload": json.dumps(payload, ensure_ascii=False)},
            files=uploads,
        )
        replay = client.post(
            f"/api/projects/{harness.project_code.lower()}/purchase-orders/"
            f"{order['id']}/supplier-invoices",
            headers={"Idempotency-Key": key},
            data={"payload": json.dumps(payload, ensure_ascii=False)},
            files=uploads,
        )

    assert created.status_code == replay.status_code == 201, created.text
    assert replay.json() == created.json()
    version_ids = created.json()["document_version_ids"]
    assert len(version_ids) == 2
    with harness.connection() as connection:
        rows = connection.execute(
            """
            SELECT versions.original_filename, versions.managed_filename,
                   versions.stored_relative_path, documents.category,
                   documents.logical_name
            FROM document_versions AS versions
            JOIN documents ON documents.id = versions.document_id
            WHERE versions.id IN (?, ?)
            ORDER BY versions.id
            """,
            version_ids,
        ).fetchall()
        assert connection.execute(
            "SELECT COUNT(*) FROM supplier_invoices"
        ).fetchone()[0] == 1
    assert [row["managed_filename"] for row in rows] == [
        "P-2026-001_进项发票_PUR-INV-001_01.pdf",
        "P-2026-001_进项发票_PUR-INV-001_02.png",
    ]
    assert [row["original_filename"] for row in rows] == [
        "供应商原票.pdf",
        "抵扣联.png",
    ]
    assert all(row["category"] == "invoice" for row in rows)
    assert len({row["logical_name"] for row in rows}) == 2
    for row in rows:
        stored = harness.settings.data_dir / row["stored_relative_path"]
        assert stored.is_file()
        assert stored.name == row["managed_filename"]


def test_supplier_invoice_multipart_invalid_payload_cleans_staged_file(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
        line_id = order["lines"][0]["id"]
        response = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/"
            f"{order['id']}/supplier-invoices",
            headers={
                "Idempotency-Key": "73100000-0000-4000-8000-000000000001"
            },
            data={
                "payload": json.dumps(
                    {
                        "invoice_no": "PUR-INV-INVALID",
                        "invoiced_on": "2026-08-31",
                        "amount_cents": -1,
                        "allocations": [
                            {
                                "purchase_order_line_id": line_id,
                                "amount_cents": 5000,
                            }
                        ],
                        "document_version_ids": [],
                    }
                )
            },
            files={"files": ("valid.pdf", b"valid", "application/pdf")},
        )

    assert response.status_code == 422
    assert not list((harness.settings.data_dir / "Temp").glob(".upload-*.tmp"))


def test_supplier_invoice_rejects_repeated_idempotency_header(tmp_path: Path) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
        line_id = order["lines"][0]["id"]
        response = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/"
            f"{order['id']}/supplier-invoices",
            headers=[
                ("Idempotency-Key", "73200000-0000-4000-8000-000000000001"),
                ("Idempotency-Key", "73200000-0000-4000-8000-000000000001"),
            ],
            json={
                "invoice_no": "PUR-INV-HEADER",
                "invoiced_on": "2026-08-31",
                "amount_cents": 5000,
                "allocations": [
                    {"purchase_order_line_id": line_id, "amount_cents": 5000}
                ],
                "document_version_ids": [],
            },
        )

    assert response.status_code == 422
    assert response.json()["error_code"] == "VALIDATION_ERROR"


def test_supplier_invoice_json_and_zero_file_multipart_replay_same_response(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
        line_id = order["lines"][0]["id"]
        payload = {
            "invoice_no": "PUR-INV-ZERO",
            "invoiced_on": "2026-08-31",
            "amount_cents": 5000,
            "allocations": [
                {"purchase_order_line_id": line_id, "amount_cents": 5000}
            ],
            "document_version_ids": [],
        }
        key = "73300000-0000-4000-8000-000000000001"
        endpoint = (
            f"/api/projects/{harness.project_code}/purchase-orders/"
            f"{order['id']}/supplier-invoices"
        )
        created = client.post(endpoint, headers={"Idempotency-Key": key}, json=payload)
        replay = client.post(
            endpoint,
            headers={"Idempotency-Key": key},
            files={
                "payload": (
                    None,
                    json.dumps(payload, ensure_ascii=False),
                    "application/json",
                )
            },
        )

    assert created.status_code == replay.status_code == 201, replay.text
    assert replay.json() == created.json()
    with harness.connection() as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM supplier_invoices"
        ).fetchone()[0] == 1


def test_supplier_invoice_multipart_database_failure_rolls_back_everything(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
    with harness.connection() as connection:
        connection.execute(
            """
            CREATE TRIGGER reject_supplier_attachment_version
            BEFORE INSERT ON document_versions
            BEGIN
                SELECT RAISE(ABORT, 'injected attachment database failure');
            END
            """
        )
    payload = {
        "invoice_no": "PUR-INV-DB-FAIL",
        "invoiced_on": "2026-08-31",
        "amount_cents": 5000,
        "allocations": [
            {
                "purchase_order_line_id": order["lines"][0]["id"],
                "amount_cents": 5000,
            }
        ],
        "document_version_ids": [],
    }
    with harness.client(raise_server_exceptions=False) as client:
        response = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/"
            f"{order['id']}/supplier-invoices",
            headers={
                "Idempotency-Key": "73400000-0000-4000-8000-000000000001"
            },
            data={"payload": json.dumps(payload)},
            files={"files": ("invoice.pdf", b"invoice", "application/pdf")},
        )

    assert response.status_code == 500
    with harness.connection() as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM supplier_invoices"
        ).fetchone()[0] == 0
        assert connection.execute("SELECT COUNT(*) FROM documents").fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM document_versions"
        ).fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM supplier_invoice_documents"
        ).fetchone()[0] == 0
    assert not [
        path
        for path in (harness.settings.data_dir / "Projects").rglob("*")
        if path.is_file()
    ]
    assert not list((harness.settings.data_dir / "Temp").glob(".upload-*.tmp"))


def test_supplier_invoice_second_publish_failure_rolls_back_everything(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
    original_publish = files.publish_staged_version
    calls = 0

    def fail_second_publish(*args: object, **kwargs: object):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise OSError("injected publish failure")
        return original_publish(*args, **kwargs)

    monkeypatch.setattr(files, "publish_staged_version", fail_second_publish)
    payload = {
        "invoice_no": "PUR-INV-FS-FAIL",
        "invoiced_on": "2026-08-31",
        "amount_cents": 5000,
        "allocations": [
            {
                "purchase_order_line_id": order["lines"][0]["id"],
                "amount_cents": 5000,
            }
        ],
        "document_version_ids": [],
    }
    with harness.client(raise_server_exceptions=False) as client:
        response = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/"
            f"{order['id']}/supplier-invoices",
            headers={
                "Idempotency-Key": "73500000-0000-4000-8000-000000000001"
            },
            data={"payload": json.dumps(payload)},
            files=[
                ("files", ("first.pdf", b"first", "application/pdf")),
                ("files", ("second.pdf", b"second", "application/pdf")),
            ],
        )

    assert response.status_code == 500
    with harness.connection() as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM supplier_invoices"
        ).fetchone()[0] == 0
        assert connection.execute("SELECT COUNT(*) FROM documents").fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM document_versions"
        ).fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM supplier_invoice_documents"
        ).fetchone()[0] == 0
    assert not [
        path
        for path in (harness.settings.data_dir / "Projects").rglob("*")
        if path.is_file()
    ]
    assert not list((harness.settings.data_dir / "Temp").glob(".upload-*.tmp"))


def test_authentication_replay_archive_and_cross_project_ordering(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
        line_id = order["lines"][0]["id"]
        endpoint = (
            f"/api/projects/{harness.project_code.lower()}/purchase-orders/"
            f"{order['id']}/supplier-payments"
        )
        headers = {"Idempotency-Key": "52000000-0000-4000-8000-000000000001"}
        payload = {
            "paid_on": "2026-08-31",
            "amount_cents": 1000,
            "payment_method": "银行转账",
            "reference_no": None,
            "allocations": [
                {"purchase_order_line_id": line_id, "amount_cents": 1000}
            ],
            "notes": None,
        }
        with TestClient(harness.app) as anonymous:
            unauthenticated = anonymous.post(
                endpoint,
                headers=headers,
                content=b"{",
            )
        assert unauthenticated.status_code == 401

        created = client.post(endpoint, headers=headers, json=payload)
        assert created.status_code == 201, created.text
        with harness.connection() as connection:
            connection.execute(
                """
                UPDATE projects
                SET status = 'archived', archived_at = ?, archive_reason = ?
                WHERE project_code_key = ?
                """,
                (
                    NOW.isoformat(),
                    "测试归档",
                    project_code_identity(harness.project_code),
                ),
            )

        replay = client.post(
            endpoint.replace(harness.project_code.lower(), harness.project_code),
            headers=headers,
            json=payload,
        )
        assert replay.status_code == 201
        assert replay.json() == created.json()
        key_reuse = client.post(
            endpoint,
            headers=headers,
            json={**payload, "reference_no": "另一个流水号"},
        )
        assert key_reuse.status_code == 409
        assert key_reuse.json()["error_code"] == "IDEMPOTENCY_KEY_REUSED"
        blocked = client.post(
            endpoint,
            headers={"Idempotency-Key": "52000000-0000-4000-8000-000000000002"},
            json=payload,
        )
        assert blocked.status_code == 409
        assert blocked.json()["error_code"] == "PROJECT_ARCHIVED"

        cross_project = client.post(
            f"/api/projects/{harness.other_project_code}/supplier-payments/"
            f"{created.json()['id']}/reverse",
            headers={"Idempotency-Key": "52000000-0000-4000-8000-000000000003"},
            json={"reason": "错误项目", "expected_revision": created.json()["revision"]},
        )
        assert cross_project.status_code == 404


def test_goods_receipt_reverse_atomically_restores_order_and_inventory(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
        order_line_id = order["lines"][0]["id"]
        receipt = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/goods-receipts",
            headers={"Idempotency-Key": "60000000-0000-4000-8000-000000000001"},
            json={
                "received_on": "2026-08-31",
                "warehouse_name": "主仓",
                "lines": [{"purchase_order_line_id": order_line_id, "quantity": "2.000"}],
                "notes": None,
            },
        )
        assert receipt.status_code == 201, receipt.text
        reverse_headers = {"Idempotency-Key": "60000000-0000-4000-8000-000000000002"}
        reverse_payload = {"reason": "供应商退货", "expected_revision": receipt.json()["revision"]}
        reversed_receipt = client.post(
            f"/api/projects/{harness.project_code.lower()}/goods-receipts/{receipt.json()['id']}/reverse",
            headers=reverse_headers,
            json=reverse_payload,
        )
        replay = client.post(
            f"/api/projects/{harness.project_code}/goods-receipts/{receipt.json()['id']}/reverse",
            headers=reverse_headers,
            json=reverse_payload,
        )
        assert reversed_receipt.status_code == replay.status_code == 200
        assert replay.json() == reversed_receipt.json()
        assert reversed_receipt.json()["reversal_reason"] == "供应商退货"
        assert reversed_receipt.json()["reversed_at"] == NOW.isoformat()
        duplicate = client.post(
            f"/api/projects/{harness.project_code}/goods-receipts/{receipt.json()['id']}/reverse",
            headers={"Idempotency-Key": "60000000-0000-4000-8000-000000000003"},
            json={"reason": "重复", "expected_revision": reversed_receipt.json()["revision"]},
        )
        assert duplicate.status_code == 409
        assert duplicate.json()["error_code"] == "GOODS_RECEIPT_ALREADY_REVERSED"

    with harness.connection() as connection:
        item = connection.execute("SELECT * FROM inventory_items").fetchone()
        assert item["quantity_milli"] == 0
        assert item["inventory_value_cents"] == 0
        order_line = connection.execute(
            "SELECT * FROM purchase_order_lines WHERE id = ?", (order_line_id,)
        ).fetchone()
        assert order_line["received_quantity_milli"] == 0
        order_row = connection.execute(
            "SELECT * FROM purchase_orders WHERE id = ?", (order["id"],)
        ).fetchone()
        assert order_row["status"] == "confirmed"
        movements = connection.execute(
            "SELECT * FROM inventory_movements ORDER BY id"
        ).fetchall()
        assert [row["quantity_delta_milli"] for row in movements] == [2000, -2000]
        assert [row["value_delta_cents"] for row in movements] == [2000, -2000]


def test_purchase_order_detail_includes_material_facts_for_each_receipt_line(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
        receipt = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/goods-receipts",
            headers={"Idempotency-Key": "61000000-0000-4000-8000-000000000001"},
            json={
                "received_on": "2026-08-31",
                "warehouse_name": "主仓",
                "lines": [
                    {
                        "purchase_order_line_id": order["lines"][0]["id"],
                        "quantity": "2.000",
                    }
                ],
                "notes": None,
            },
        )
        assert receipt.status_code == 201, receipt.text
        detail = client.get(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}"
        )

    assert detail.status_code == 200
    receipt_fact = detail.json()["goods_receipts"][0]
    assert receipt_fact["warehouse_name"] == "主仓"
    assert receipt_fact["status"] == "active"
    assert receipt_fact["lines"] == [
        {
            **receipt.json()["lines"][0],
            "material_name": "接触器",
            "material_model": "LC1D09",
            "unit": "PCS",
        }
    ]


def test_goods_receipt_reverse_rolls_back_when_inventory_would_be_negative(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
        receipt = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/goods-receipts",
            headers={"Idempotency-Key": "61000000-0000-4000-8000-000000000001"},
            json={
                "received_on": "2026-08-31",
                "warehouse_name": "主仓",
                "lines": [{"purchase_order_line_id": order["lines"][0]["id"], "quantity": "2.000"}],
                "notes": None,
            },
        )
        assert receipt.status_code == 201
        with harness.connection() as connection:
            connection.execute(
                "UPDATE inventory_items SET quantity_milli = 1000, inventory_value_cents = 1000"
            )
        reverse = client.post(
            f"/api/projects/{harness.project_code}/goods-receipts/{receipt.json()['id']}/reverse",
            headers={"Idempotency-Key": "61000000-0000-4000-8000-000000000002"},
            json={"reason": "供应商退货", "expected_revision": receipt.json()["revision"]},
        )
    assert reverse.status_code == 409
    assert reverse.json()["error_code"] == "RECEIPT_REVERSAL_INSUFFICIENT_INVENTORY"
    with harness.connection() as connection:
        assert connection.execute("SELECT status FROM goods_receipts").fetchone()[0] == "active"
        assert connection.execute("SELECT COUNT(*) FROM inventory_movements").fetchone()[0] == 1


def test_order_and_overview_reload_active_totals_and_persist_reversal_audit(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        order = _create_confirmed_order(client, harness)
        line_id = order["lines"][0]["id"]
        payment = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/supplier-payments",
            headers={"Idempotency-Key": "80000000-0000-4000-8000-000000000001"},
            json={
                "paid_on": "2026-08-31",
                "amount_cents": 1000,
                "payment_method": "银行转账",
                "reference_no": "PAY-AUDIT",
                "allocations": [
                    {"purchase_order_line_id": line_id, "amount_cents": 1000}
                ],
                "notes": None,
            },
        )
        assert payment.status_code == 201, payment.text
        invoice = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/supplier-invoices",
            headers={"Idempotency-Key": "80000000-0000-4000-8000-000000000002"},
            json={
                "invoice_no": "INV-AUDIT",
                "invoiced_on": "2026-08-31",
                "amount_cents": 1500,
                "allocations": [
                    {"purchase_order_line_id": line_id, "amount_cents": 1500}
                ],
                "document_version_ids": [harness.document_version_id],
            },
        )
        assert invoice.status_code == 201, invoice.text
        receipt = client.post(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}/goods-receipts",
            headers={"Idempotency-Key": "80000000-0000-4000-8000-000000000003"},
            json={
                "received_on": "2026-08-31",
                "warehouse_name": "主仓",
                "lines": [
                    {"purchase_order_line_id": line_id, "quantity": "2.000"}
                ],
                "notes": None,
            },
        )
        assert receipt.status_code == 201, receipt.text

        active_detail = client.get(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}"
        )
        assert active_detail.status_code == 200
        assert active_detail.json()["paid_amount_cents"] == 1000
        assert active_detail.json()["invoiced_amount_cents"] == 1500
        assert active_detail.json()["received_amount_cents"] == 2000
        assert active_detail.json()["supplier_payments"] == [payment.json()]
        assert active_detail.json()["supplier_invoices"] == [invoice.json()]
        assert active_detail.json()["goods_receipts"] == [receipt.json()]
        active_overview = client.get(
            f"/api/projects/{harness.project_code}/procurement-overview"
        )
        assert active_overview.json()["procurement_paid_cents"] == 1000
        assert active_overview.json()["procurement_invoiced_cents"] == 1500
        assert active_overview.json()["procurement_received_cents"] == 2000

        reversed_payment = client.post(
            f"/api/projects/{harness.project_code}/supplier-payments/{payment.json()['id']}/reverse",
            headers={"Idempotency-Key": "80000000-0000-4000-8000-000000000004"},
            json={
                "reason": "付款退回",
                "expected_revision": payment.json()["revision"],
            },
        )
        reversed_invoice = client.post(
            f"/api/projects/{harness.project_code}/supplier-invoices/{invoice.json()['id']}/reverse",
            headers={"Idempotency-Key": "80000000-0000-4000-8000-000000000005"},
            json={
                "reason": "发票作废",
                "expected_revision": invoice.json()["revision"],
            },
        )
        reversed_receipt = client.post(
            f"/api/projects/{harness.project_code}/goods-receipts/{receipt.json()['id']}/reverse",
            headers={"Idempotency-Key": "80000000-0000-4000-8000-000000000006"},
            json={
                "reason": "到货退回",
                "expected_revision": receipt.json()["revision"],
            },
        )
        assert reversed_payment.status_code == 200
        assert reversed_invoice.status_code == 200
        assert reversed_receipt.status_code == 200

    with harness.client() as restarted_client:
        reloaded_detail = restarted_client.get(
            f"/api/projects/{harness.project_code}/purchase-orders/{order['id']}"
        )
        reloaded_overview = restarted_client.get(
            f"/api/projects/{harness.project_code}/procurement-overview"
        )
    assert reloaded_detail.status_code == 200
    assert reloaded_detail.json()["paid_amount_cents"] == 0
    assert reloaded_detail.json()["invoiced_amount_cents"] == 0
    assert reloaded_detail.json()["received_amount_cents"] == 0
    assert reloaded_detail.json()["supplier_payments"] == [reversed_payment.json()]
    assert reloaded_detail.json()["supplier_invoices"] == [reversed_invoice.json()]
    assert reloaded_detail.json()["goods_receipts"] == [reversed_receipt.json()]
    assert reloaded_overview.json()["procurement_paid_cents"] == 0
    assert reloaded_overview.json()["procurement_invoiced_cents"] == 0
    assert reloaded_overview.json()["procurement_received_cents"] == 0
    with harness.connection() as connection:
        assert connection.execute(
            """
            SELECT document_version_id FROM supplier_invoice_documents
            WHERE supplier_invoice_id = ?
            """,
            (invoice.json()["id"],),
        ).fetchone()[0] == harness.document_version_id
        assert dict(
            connection.execute(
                "SELECT reversal_reason, reversed_at FROM supplier_payments"
            ).fetchone()
        ) == {"reversal_reason": "付款退回", "reversed_at": NOW.isoformat()}
        assert dict(
            connection.execute(
                "SELECT reversal_reason, reversed_at FROM supplier_invoices"
            ).fetchone()
        ) == {"reversal_reason": "发票作废", "reversed_at": NOW.isoformat()}
        assert dict(
            connection.execute(
                "SELECT reversal_reason, reversed_at FROM goods_receipts"
            ).fetchone()
        ) == {"reversal_reason": "到货退回", "reversed_at": NOW.isoformat()}


def test_quote_export_uses_customer_only_dto_without_cost_or_hidden_content(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        procurement_list = _create_confirmed_list(client, harness.project_code)
        headers = {"Idempotency-Key": "70000000-0000-4000-8000-000000000001"}
        payload = {
            "title": " 项目报价单 ",
            "customer_company_id": harness.customer_company_id,
            "notes": "有效期 30 天",
        }
        created = client.post(
            f"/api/projects/{harness.project_code.lower()}/procurement-lists/{procurement_list['id']}/quote-exports",
            headers=headers,
            json=payload,
        )
        replay = client.post(
            f"/api/projects/{harness.project_code}/procurement-lists/{procurement_list['id']}/quote-exports",
            headers=headers,
            json=payload,
        )
        assert created.status_code == replay.status_code == 201
        assert replay.json() == created.json()
        cross_project = client.post(
            f"/api/projects/{harness.other_project_code}/procurement-lists/"
            f"{procurement_list['id']}/quote-exports",
            headers={"Idempotency-Key": "70000000-0000-4000-8000-000000000002"},
            json=payload,
        )
        assert cross_project.status_code == 404
        download = client.get(created.json()["download_url"])
        assert download.status_code == 200

    workbook = load_workbook(BytesIO(download.content), data_only=False)
    assert workbook.sheetnames == ["报价单"]
    worksheet = workbook["报价单"]
    assert worksheet.sheet_state == "visible"
    assert all(dimension.hidden is not True for dimension in worksheet.column_dimensions.values())
    assert worksheet["A1"].value == "标题"
    assert worksheet["B1"].value == "项目报价单"
    assert worksheet["A2"].value == "客户公司"
    assert worksheet["B2"].value == "客户公司"
    assert worksheet["A3"].value == "备注"
    assert worksheet["B3"].value == "有效期 30 天"
    assert [cell.value for cell in worksheet[5]] == [
        "序号",
        "大类",
        "名称",
        "规格",
        "品牌",
        "型号",
        "数量",
        "单位",
        "报价单价（元）",
        "报价金额（元）",
    ]
    assert [cell.value for cell in worksheet[6]] == [
        1,
        "电气",
        "接触器",
        "AC220V",
        "施耐德",
        "LC1D09",
        "5.000",
        "PCS",
        "15.00",
        "75.00",
    ]
    assert all(cell.data_type != "f" for row in worksheet.iter_rows() for cell in row)
    assert all(sheet.sheet_state == "visible" for sheet in workbook.worksheets)


def test_quote_export_history_is_project_scoped_and_remains_downloadable(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        procurement_list = _create_confirmed_list(client, harness.project_code)
        created_exports = []
        for index in (1, 2):
            response = client.post(
                f"/api/projects/{harness.project_code}/procurement-lists/"
                f"{procurement_list['id']}/quote-exports",
                headers={
                    "Idempotency-Key": (
                        f"71000000-0000-4000-8000-{index:012d}"
                    )
                },
                json={
                    "title": f"项目报价单 V{index}",
                    "customer_company_id": harness.customer_company_id,
                    "notes": None,
                },
            )
            assert response.status_code == 201
            created_exports.append(response.json())

        history = client.get(
            f"/api/projects/{harness.project_code.lower()}/quote-exports",
            params={"page": 1, "page_size": 1},
        )
        other_project = client.get(
            f"/api/projects/{harness.other_project_code}/quote-exports"
        )
        old_download = client.get(created_exports[0]["download_url"])

    assert history.status_code == 200
    assert history.json()["total"] == 2
    assert history.json()["page"] == 1
    assert history.json()["page_size"] == 1
    assert [item["title"] for item in history.json()["items"]] == [
        "项目报价单 V2"
    ]
    assert other_project.status_code == 200
    assert other_project.json()["items"] == []
    assert old_download.status_code == 200
    assert old_download.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_quote_export_rejects_company_that_is_not_the_project_customer(
    tmp_path: Path,
) -> None:
    harness = _build_harness(tmp_path)
    with harness.client() as client:
        procurement_list = _create_confirmed_list(client, harness.project_code)
        response = client.post(
            f"/api/projects/{harness.project_code}/procurement-lists/"
            f"{procurement_list['id']}/quote-exports",
            headers={"Idempotency-Key": "70000000-0000-4000-8000-000000000099"},
            json={
                "title": "项目报价单",
                "customer_company_id": harness.supplier_company_id,
                "notes": None,
            },
        )

    assert response.status_code == 409
    assert response.json()["error_code"] == "PROJECT_CUSTOMER_MISMATCH"
    assert response.json()["field_errors"] == {
        "customer_company_id": ["必须使用项目绑定的客户公司"]
    }
    with harness.connection() as connection:
        assert connection.execute("SELECT COUNT(*) FROM quote_exports").fetchone()[0] == 0
